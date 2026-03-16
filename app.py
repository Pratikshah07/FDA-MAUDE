"""
Flask application for MAUDE data processing web interface with authentication.
"""
import os
import json
import csv
import re
import time
import threading
import uuid
from datetime import datetime, timezone
from urllib.parse import urlparse, urlunparse, parse_qs, urlencode
import requests
from flask import Flask, request, render_template, send_file, jsonify, redirect, url_for, session, Response, stream_with_context
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.utils import secure_filename
import tempfile
import io

from backend.auth import FirebaseAuthManager, User
from backend.txt_to_csv_converter import TxtToCsvConverter, get_txt_preview
from backend.csv_viewer import LargeCSVViewer, get_csv_page, get_csv_info
from config import GROQ_API_KEY, SECRET_KEY, FIREBASE_CONFIG

# Heavy processing modules are imported lazily inside route functions so the
# app binds quickly on startup (critical for Render cold-start wake-up time).
# Python caches modules in sys.modules, so the import only pays the cost once.

# Default IMDRF Annexure file bundled with the project
DEFAULT_IMDRF_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'Annexes A-G consolidated.xlsx')

app = Flask(__name__)
app.config['SECRET_KEY'] = SECRET_KEY
# Note: MAX_CONTENT_LENGTH is set to None to allow streaming uploads of 10GB+ files
# Validation is done at the application level for specific routes
app.config['MAX_CONTENT_LENGTH'] = None
# Use /tmp for Vercel (serverless) or temp directory for local
if os.path.exists('/tmp'):
    app.config['UPLOAD_FOLDER'] = os.path.join('/tmp', 'maude_uploads')
elif os.getenv('TMPDIR'):
    app.config['UPLOAD_FOLDER'] = os.path.join(os.getenv('TMPDIR'), 'maude_uploads')
elif os.getenv('TMP'):
    app.config['UPLOAD_FOLDER'] = os.path.join(os.getenv('TMP'), 'maude_uploads')
else:
    app.config['UPLOAD_FOLDER'] = os.path.join(tempfile.gettempdir(), 'maude_uploads')
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], 'jobs'), exist_ok=True)

# In-memory job store for async processing
PROCESS_JOBS = {}
PROCESS_JOBS_LOCK = threading.Lock()
MAUDE_EXPORT_JOBS = {}
MAUDE_EXPORT_LOCK = threading.Lock()
PIPELINE_JOBS = {}
PIPELINE_JOBS_LOCK = threading.Lock()

# Firebase config (available to all templates via context processor)
app.config['FIREBASE_CONFIG'] = FIREBASE_CONFIG

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'info'


@login_manager.unauthorized_handler
def _handle_unauthorized():
    """Return JSON for API routes, otherwise redirect to login."""
    try:
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Authentication required'}), 401
    except Exception:
        pass
    return redirect(url_for('login', next=request.path))

# Initialize Firebase Auth Manager (uses PyJWT + Google public keys, no service account needed)
auth_manager = FirebaseAuthManager()

ALLOWED_EXTENSIONS = {'csv', 'xlsx', 'xls'}


@login_manager.user_loader
def load_user(user_id):
    """Load user from Flask session for Flask-Login."""
    user_data = session.get('firebase_user')
    if user_data and user_data.get('uid') == user_id:
        return FirebaseAuthManager.get_user_from_session(user_data)
    return None


@app.context_processor
def inject_firebase_config():
    """Make firebase_config available in all templates."""
    return dict(firebase_config=app.config.get('FIREBASE_CONFIG', {}))


def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def _set_job_status(job_id, **updates):
    with PROCESS_JOBS_LOCK:
        job = PROCESS_JOBS.setdefault(job_id, {})
        job.update(updates)
        _save_job_file('process', job_id, job)


def _get_job(job_id):
    with PROCESS_JOBS_LOCK:
        job = PROCESS_JOBS.get(job_id)
        if job:
            return job
    return _load_job_file('process', job_id)


def _set_export_status(job_id, **updates):
    with MAUDE_EXPORT_LOCK:
        job = MAUDE_EXPORT_JOBS.setdefault(job_id, {})
        if 'processed' in updates:
            existing = job.get('processed') or 0
            updates['processed'] = max(existing, updates.get('processed') or 0)
        if 'scanned' in updates:
            existing_scanned = job.get('scanned') or 0
            updates['scanned'] = max(existing_scanned, updates.get('scanned') or 0)
        job.update(updates)
        _save_job_file('export', job_id, job)


def _get_export_job(job_id):
    with MAUDE_EXPORT_LOCK:
        job = MAUDE_EXPORT_JOBS.get(job_id)
        if job:
            return job
    return _load_job_file('export', job_id)


def _set_pipeline_status(job_id, **updates):
    with PIPELINE_JOBS_LOCK:
        job = PIPELINE_JOBS.setdefault(job_id, {})
        if 'processed' in updates:
            existing = job.get('processed') or 0
            updates['processed'] = max(existing, updates.get('processed') or 0)
        if 'scanned' in updates:
            existing_scanned = job.get('scanned') or 0
            updates['scanned'] = max(existing_scanned, updates.get('scanned') or 0)
        job.update(updates)
        _save_job_file('pipeline', job_id, job)


def _get_pipeline_job(job_id):
    with PIPELINE_JOBS_LOCK:
        job = PIPELINE_JOBS.get(job_id)
        if job:
            return job
    return _load_job_file('pipeline', job_id)


def _job_file_path(prefix: str, job_id: str) -> str:
    jobs_dir = os.path.join(app.config['UPLOAD_FOLDER'], 'jobs')
    return os.path.join(jobs_dir, f"{prefix}_{job_id}.json")


def _save_job_file(prefix: str, job_id: str, payload: dict) -> None:
    try:
        path = _job_file_path(prefix, job_id)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(payload, f)
    except Exception:
        pass


def _load_job_file(prefix: str, job_id: str):
    try:
        path = _job_file_path(prefix, job_id)
        if not os.path.exists(path):
            return None
        with open(path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return None


@app.route('/api/maude/export/status/<job_id>', methods=['GET'])
@login_required
def maude_export_status(job_id):
    job = _get_export_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'status': job.get('status'),
        'error': job.get('error'),
        'processed': job.get('processed'),
        'total': job.get('total'),
        'scanned': job.get('scanned')
    }), 200


@app.route('/api/maude/export/download/<job_id>', methods=['GET'])
@login_required
def maude_export_download(job_id):
    job = _get_export_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'Job not completed'}), 400
    output_path = job.get('output_path')
    output_filename = job.get('output_filename')
    if not output_path or not os.path.exists(output_path):
        return jsonify({'error': 'Output file not found'}), 404
    return send_file(
        output_path,
        as_attachment=True,
        download_name=output_filename,
        mimetype='text/csv'
    )


def _run_processing_job(job_id, input_path, output_path, output_filename, imdrf_path, user_id):
    _set_job_status(job_id, status='running')
    try:
        from backend.processor import MAUDEProcessor  # lazy import
        processor = MAUDEProcessor()

        # Use uploaded IMDRF file, or fall back to bundled default
        annex_path = imdrf_path if (imdrf_path and os.path.exists(imdrf_path)) else DEFAULT_IMDRF_PATH
        if os.path.exists(annex_path):
            processor.load_imdrf_structure(annex_path)
        else:
            print("WARNING: No IMDRF annex file found. IMDRF codes will be blank.")

        stats = processor.process_file(input_path, output_path)

        critical_failures = []
        if not stats['validation'].get('column_count_correct', False):
            critical_failures.append('Column count check failed')
        if not stats['validation'].get('file_will_open', False):
            critical_failures.append('File integrity check failed')
        if not stats['validation'].get('date_format_correct', False):
            critical_failures.append('Date format check failed (HARD STOP: no literal "nan", all dates must be DD-MM-YYYY)')
        if not stats['validation'].get('imdrf_adjacent', False):
            critical_failures.append('IMDRF Code column position check failed (HARD STOP: must be adjacent to Device Problem)')
        if not stats['validation'].get('imdrf_codes_valid', False):
            critical_failures.append('IMDRF codes validation failed (HARD STOP: all codes must exist in Annex)')

        warnings = []
        if not stats['validation'].get('no_timestamps', False):
            warnings.append('Timestamp check failed (non-critical)')

        if critical_failures:
            error_msg = f"Validation failed (HARD STOP): {', '.join(critical_failures)}"
            if warnings:
                error_msg += f" | Warnings: {', '.join(warnings)}"
            _set_job_status(job_id, status='failed', error=error_msg)
            return

        _set_job_status(
            job_id,
            status='done',
            output_path=output_path,
            output_filename=output_filename,
            user_id=user_id
        )
    except Exception as e:
        try:
            app.logger.exception("Processing failed")
        except Exception:
            pass
        _set_job_status(job_id, status='failed', error=f"Processing failed: {str(e)}")


@app.route('/process/status/<job_id>', methods=['GET'])
@login_required
def process_status(job_id):
    job = _get_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({'status': job.get('status'), 'error': job.get('error')}), 200


@app.route('/process/download/<job_id>', methods=['GET'])
@login_required
def process_download(job_id):
    job = _get_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'Job not completed'}), 400
    output_path = job.get('output_path')
    output_filename = job.get('output_filename')
    if not output_path or not os.path.exists(output_path):
        return jsonify({'error': 'Output file not found'}), 404
    return send_file(
        output_path,
        as_attachment=True,
        download_name=output_filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/health')
def health_check():
    """Lightweight health check endpoint for Render.com.
    Returns plain-text 200 immediately — no template rendering, no DB calls.
    Render uses this to determine whether the service is ready.
    """
    return 'ok', 200, {'Content-Type': 'text/plain'}


# Authentication Routes
@app.route('/login')
def login():
    """Render login page."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/register')
def register():
    """Render registration page."""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    return render_template('register.html')


@app.route('/forgot-password')
def forgot_password():
    """Render forgot password page."""
    return render_template('forgot_password.html')


@app.route('/logout')
@login_required
def logout():
    """Logout user and clear Firebase session."""
    logout_user()
    session.pop('firebase_user', None)
    return redirect(url_for('login'))


# API Routes
@app.route('/api/session-login', methods=['POST'])
def api_session_login():
    """Verify Firebase ID token and create Flask session."""
    data = request.get_json()
    id_token = data.get('idToken', '')

    if not id_token:
        return jsonify({'error': 'ID token is required'}), 400

    user = auth_manager.verify_id_token(id_token)
    if not user:
        return jsonify({'error': 'Invalid or expired token'}), 401

    # Store user data in Flask session
    session['firebase_user'] = {
        'uid': user.id,
        'email': user.email,
        'name': user.name,
    }

    # Log in with Flask-Login (sets the session cookie)
    login_user(user, remember=True)

    return jsonify({
        'success': True,
        'message': 'Session created',
        'user': {
            'email': user.email,
            'name': user.name,
        }
    }), 200


# Protected Routes
@app.route('/')
@login_required
def index():
    """Render main upload page."""
    return render_template('index.html', user=current_user)


@app.route('/process', methods=['POST'])
@login_required
def process_file():
    """Process uploaded MAUDE file."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file type. Please upload CSV or Excel file.'}), 400
    
    # Note: GROQ_API_KEY is optional - app works without it using deterministic methods only
    # if not GROQ_API_KEY:
    #     return jsonify({'error': 'GROQ_API_KEY not configured. Please set environment variable.'}), 500
    
    try:
        # Save uploaded file
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")
        file.save(input_path)
        
        # Create output file path
        output_filename = f"cleaned_{filename.rsplit('.', 1)[0]}.xlsx"
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
        
        imdrf_path = None
        if 'imdrf_file' in request.files:
            imdrf_file = request.files['imdrf_file']
            if imdrf_file.filename and imdrf_file.filename.strip():
                imdrf_path = os.path.join(app.config['UPLOAD_FOLDER'], f"imdrf_{secure_filename(imdrf_file.filename)}")
                imdrf_file.save(imdrf_path)

        async_requested = request.form.get('async') == '1'
        if async_requested:
            job_id = str(uuid.uuid4())
            _set_job_status(job_id, status='queued', user_id=current_user.id)
            thread = threading.Thread(
                target=_run_processing_job,
                args=(job_id, input_path, output_path, output_filename, imdrf_path, current_user.id),
                daemon=True
            )
            thread.start()
            return jsonify({
                'job_id': job_id,
                'status_url': url_for('process_status', job_id=job_id),
                'download_url': url_for('process_download', job_id=job_id)
            }), 202

        # Sync processing (fallback)
        from backend.processor import MAUDEProcessor  # lazy import
        processor = MAUDEProcessor()

        # Use uploaded IMDRF file, or fall back to bundled default
        annex_path = imdrf_path if (imdrf_path and os.path.exists(imdrf_path)) else DEFAULT_IMDRF_PATH
        if os.path.exists(annex_path):
            processor.load_imdrf_structure(annex_path)
        else:
            print("WARNING: No IMDRF annex file found. IMDRF codes will be blank.")

        stats = processor.process_file(input_path, output_path)
        
        # Check validation - HARD STOPS for critical failures
        critical_failures = []
        if not stats['validation'].get('column_count_correct', False):
            critical_failures.append('Column count check failed')
        if not stats['validation'].get('file_will_open', False):
            critical_failures.append('File integrity check failed')
        if not stats['validation'].get('date_format_correct', False):
            critical_failures.append('Date format check failed (HARD STOP: no literal "nan", all dates must be DD-MM-YYYY)')
        if not stats['validation'].get('imdrf_adjacent', False):
            critical_failures.append('IMDRF Code column position check failed (HARD STOP: must be adjacent to Device Problem)')
        if not stats['validation'].get('imdrf_codes_valid', False):
            critical_failures.append('IMDRF codes validation failed (HARD STOP: all codes must exist in Annex)')
        
        # Non-critical validations (warn but don't fail)
        warnings = []
        if not stats['validation'].get('no_timestamps', False):
            warnings.append('Timestamp check failed (non-critical)')
        
        # HARD STOP on critical failures
        if critical_failures:
            error_msg = f"Validation failed (HARD STOP): {', '.join(critical_failures)}"
            if warnings:
                error_msg += f" | Warnings: {', '.join(warnings)}"
            
            return jsonify({
                'error': error_msg,
                'validation_results': stats['validation'],
                'stats': stats,
                'failed_checks': critical_failures,
                'warnings': warnings
            }), 400
        elif warnings:
            # Log warnings but don't fail
            print(f"Validation warnings: {', '.join(warnings)}")
        
        # Return file for download
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        try:
            import traceback
            app.logger.exception("Processing failed")
            traceback.print_exc()
        except Exception:
            pass
        return jsonify({'error': f'Processing failed: {str(e)}'}), 500


# IMDRF Insights Routes
@app.route('/imdrf-insights')
@login_required
def imdrf_insights_page():
    """Render IMDRF Insights page."""
    return render_template('imdrf_insights.html', user=current_user)


@app.route('/api/imdrf-insights/prepare', methods=['POST'])
@login_required
def api_prepare_insights():
    """Prepare uploaded CSV or Excel file for IMDRF insights analysis."""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400

    # Check file extension
    allowed_extensions = {'.csv', '.xlsx', '.xls'}
    file_ext = os.path.splitext(file.filename.lower())[1]
    if file_ext not in allowed_extensions:
        return jsonify({'error': 'Invalid file type. Please upload CSV, XLS, or XLSX file.'}), 400

    # Get level from form data (default to 1 for backward compatibility)
    level = int(request.form.get('level', 1))
    if level not in [1, 2, 3]:
        return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

    try:
        # Save uploaded file with unique name
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")
        file.save(input_path)

        from backend.imdrf_insights import prepare_data_for_insights  # lazy import
        # Prepare data for insights at the specified level
        result = prepare_data_for_insights(input_path, level=level, annex_file_path=DEFAULT_IMDRF_PATH)

        # Validate minimum 1-year date range
        valid_dates = result['df_exploded']['parsed_date'].dropna()
        if len(valid_dates) > 0:
            date_span = (valid_dates.max() - valid_dates.min()).days
            if date_span < 365:
                os.remove(input_path)
                return jsonify({
                    'error': f'Uploaded file only contains {date_span} days of data '
                             f'({valid_dates.min().strftime("%Y-%m-%d")} to {valid_dates.max().strftime("%Y-%m-%d")}). '
                             f'At least 1 year (365 days) is required for meaningful trend analysis.'
                }), 400

        # Store file info in session for later use (level can be toggled later)
        session[f'insights_file_{file_id}'] = {
            'path': input_path
        }

        return jsonify({
            'success': True,
            'file_id': file_id,
            'all_prefixes': result['all_prefixes'],
            'all_manufacturers': result['all_manufacturers'],
            'prefix_counts': result.get('prefix_counts', {}),
            'total_rows': result['total_rows'],
            'rows_with_imdrf': result['rows_with_imdrf'],
            'rows_with_dates': result['rows_with_dates'],
            'level': level,
            'level_label': result.get('level_label', f'Level-{level}')
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/prepare-from-pipeline', methods=['POST'])
@login_required
def api_prepare_insights_from_pipeline():
    """Prepare IMDRF insights using a cleaned file produced by the pipeline."""
    data = request.get_json() or {}
    job_id = data.get('job_id')
    level = int(data.get('level', 1))

    if not job_id:
        return jsonify({'error': 'Missing job_id'}), 400
    if level not in [1, 2, 3]:
        return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

    job = _get_pipeline_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'Pipeline job is not complete yet'}), 409

    cleaned_path = job.get('output_path')
    if not cleaned_path or not os.path.exists(cleaned_path):
        return jsonify({'error': 'Cleaned file not found for this job'}), 404

    try:
        from backend.imdrf_insights import prepare_data_for_insights  # lazy import
        result = prepare_data_for_insights(cleaned_path, level=level, annex_file_path=DEFAULT_IMDRF_PATH)

        # Reuse the cleaned file path as the file_id key for refresh calls
        file_id = f"{current_user.id}_{job_id}"
        session[f'insights_file_{file_id}'] = {'path': cleaned_path}

        return jsonify({
            'success': True,
            'file_id': file_id,
            'all_prefixes': result['all_prefixes'],
            'all_manufacturers': result['all_manufacturers'],
            'prefix_counts': result.get('prefix_counts', {}),
            'total_rows': result['total_rows'],
            'rows_with_imdrf': result['rows_with_imdrf'],
            'rows_with_dates': result['rows_with_dates'],
            'level': level,
            'level_label': result.get('level_label', f'Level-{level}')
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/refresh', methods=['POST'])
@login_required
def api_refresh_insights():
    """Re-prepare IMDRF insights data for a new level without re-uploading."""
    data = request.get_json() or {}

    file_id = data.get('file_id')
    level = int(data.get('level', 1))

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    if level not in [1, 2, 3]:
        return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

    try:
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        from backend.imdrf_insights import prepare_data_for_insights  # lazy import (cached after first call)
        result = prepare_data_for_insights(file_path, level=level, annex_file_path=DEFAULT_IMDRF_PATH)

        return jsonify({
            'success': True,
            'file_id': file_id,
            'all_prefixes': result['all_prefixes'],
            'all_manufacturers': result['all_manufacturers'],
            'prefix_counts': result.get('prefix_counts', {}),
            'total_rows': result['total_rows'],
            'rows_with_imdrf': result['rows_with_imdrf'],
            'rows_with_dates': result['rows_with_dates'],
            'level': level,
            'level_label': result.get('level_label', f'Level-{level}')
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/top-manufacturers', methods=['GET'])
@login_required
def api_top_manufacturers():
    """Get top manufacturers for a specific IMDRF prefix."""
    prefix = request.args.get('prefix')
    file_id = request.args.get('file_id')

    if not prefix or not file_id:
        return jsonify({'error': 'Missing prefix or file_id parameter'}), 400

    try:
        # Retrieve file info from session
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Handle both old format (string) and new format (dict)
        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        level = int(request.args.get('level', 1))
        if level not in [1, 2, 3]:
            return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        from backend.imdrf_insights import prepare_data_for_insights, get_top_manufacturers_for_prefix  # lazy import
        # Prepare data again (in memory) at the same level
        result = prepare_data_for_insights(file_path, level=level)
        df_exploded = result['df_exploded']
        mfr_col = result['mfr_col']

        # Get top manufacturers
        top_mfrs = get_top_manufacturers_for_prefix(df_exploded, prefix, mfr_col, top_n=5)

        return jsonify({
            'success': True,
            'top_manufacturers': top_mfrs
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/analyze', methods=['POST'])
@login_required
def api_analyze_insights():
    """Perform IMDRF insights analysis."""
    data = request.get_json()

    file_id = data.get('file_id')
    prefix = data.get('prefix')
    manufacturers = data.get('manufacturers', [])
    grain = data.get('grain', 'M')
    date_from = data.get('date_from')  # optional YYYY-MM-DD
    date_to = data.get('date_to')      # optional YYYY-MM-DD

    if not file_id or not prefix or not manufacturers:
        return jsonify({'error': 'Missing required parameters'}), 400

    try:
        # Retrieve file info from session
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Handle both old format (string) and new format (dict)
        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        level = int(data.get('level', 1))
        if level not in [1, 2, 3]:
            return jsonify({'error': 'Invalid level. Must be 1, 2, or 3.'}), 400

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        from backend.imdrf_insights import prepare_data_for_insights, analyze_imdrf_insights  # lazy import
        # Prepare data at the same level
        result = prepare_data_for_insights(file_path, level=level)
        df_exploded = result['df_exploded']

        # Perform analysis
        analysis_result = analyze_imdrf_insights(
            df_exploded,
            prefix,
            manufacturers,
            grain,
            level=level,
            date_from=date_from,
            date_to=date_to
        )

        # Convert pandas data to JSON-serializable format
        manufacturer_series = {}
        for mfr, series in analysis_result['manufacturer_series'].items():
            manufacturer_series[mfr] = {
                'dates': series.index.strftime('%Y-%m-%d').tolist(),
                'values': series.tolist()
            }

        response_data = {
            'success': True,
            'threshold': analysis_result['threshold'],
            'total_selected_code': analysis_result['total_selected_code'],
            'total_all_codes': analysis_result['total_all_codes'],
            'manufacturer_series': manufacturer_series,
            'date_range': analysis_result['date_range'].strftime('%Y-%m-%d').tolist() if len(analysis_result['date_range']) > 0 else [],
            'statistics': analysis_result['statistics'],
            'grain': grain,
            'selected_prefix': prefix,
            'level': level,
            'level_label': analysis_result.get('level_label', f'Level-{level}')
        }

        return jsonify(_sanitize_json(response_data)), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/annex-status', methods=['GET'])
@login_required
def api_annex_status():
    """Get the status of the IMDRF Annex file loading."""
    try:
        from backend.imdrf_annex_validator import get_annex_status  # lazy import
        status = get_annex_status()
        return jsonify({
            'success': True,
            'annex_status': status
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/download-code-counts-xlsx', methods=['GET'])
@login_required
def api_download_code_counts_xlsx():
    """Download IMDRF code counts for all levels as an XLSX file."""
    file_id = request.args.get('file_id')

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    try:
        file_info = session.get(f'insights_file_{file_id}')
        if not file_info:
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        if isinstance(file_info, str):
            file_path = file_info
        else:
            file_path = file_info.get('path')

        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        from backend.imdrf_insights import (  # lazy import
            get_imdrf_code_manufacturer_monthly_counts,
            get_patient_problem_e_code_mfr_monthly_counts,
            get_imdrf_code_counts_all_levels_with_descriptions,
            get_imdrf_code_monthly_counts,
        )
        mfr_monthly = get_imdrf_code_manufacturer_monthly_counts(file_path)
        pat_mfr_monthly = get_patient_problem_e_code_mfr_monthly_counts(file_path, DEFAULT_IMDRF_PATH)
        summary_counts = get_imdrf_code_counts_all_levels_with_descriptions(file_path, DEFAULT_IMDRF_PATH)
        summary_monthly = get_imdrf_code_monthly_counts(file_path)

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        import calendar as _calendar

        all_months = mfr_monthly['months']
        data_by_level = mfr_monthly['data']

        def _fmt_month(ym):
            try:
                y, m = ym.split('-')
                return f"{_calendar.month_abbr[int(m)]}-{y}"
            except Exception:
                return ym

        month_labels = [_fmt_month(m) for m in all_months]

        # Styles
        hdr_font   = Font(bold=True, color="FFFFFF")
        hdr_fill   = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        code_font  = Font(bold=True, color="1E40AF")
        code_fill  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        center     = Alignment(horizontal='center')

        level_sheet_names = {1: 'Level-1 (3 chars)', 2: 'Level-2 (5 chars)', 3: 'Level-3 (7 chars)'}

        wb = Workbook()
        wb.remove(wb.active)  # remove default blank sheet

        # ── Sheet 1: Summary (All Manufacturers) — one sheet, all 3 levels ──
        summary_months = summary_monthly.get('months', [])
        summary_month_labels = [_fmt_month(m) for m in summary_months]
        num_summary_months = len(summary_months)

        ws_sum = wb.create_sheet(title='Summary (All Mfrs)')
        sum_hdr_font  = Font(bold=True, color="FFFFFF")
        sum_hdr_fill  = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        sum_lvl_font  = Font(bold=True, color="1E3A5F")
        sum_lvl_fill  = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")
        sum_tot_font  = Font(bold=True)
        sum_tot_fill  = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid")
        num_sum_cols = 4 + num_summary_months  # Code + Desc + Total + Avg + months

        sum_row = 1  # explicit row counter
        for level in [1, 2, 3]:
            level_label_sum = f'LEVEL-{level} ({[3,5,7][level-1]} chars)'
            # Level header row (coloured background across all columns)
            ws_sum.cell(row=sum_row, column=1, value=level_label_sum).font = sum_lvl_font
            for c in range(1, num_sum_cols + 1):
                ws_sum.cell(row=sum_row, column=c).fill = sum_lvl_fill
            sum_row += 1

            # Column header row
            col_hdr = ['IMDRF Code', 'Description', 'Total', 'Avg/Month'] + summary_month_labels
            for ci, h in enumerate(col_hdr, 1):
                cell = ws_sum.cell(row=sum_row, column=ci, value=h)
                cell.font = sum_hdr_font
                cell.fill = sum_hdr_fill
                cell.alignment = center
            sum_row += 1

            level_codes = summary_counts.get(level, {})
            level_monthly = summary_monthly.get('counts', {}).get(level, {})
            # Sort by total descending
            for code in sorted(level_codes.keys(), key=lambda c: -level_codes[c].get('count', 0)):
                entry = level_codes[code]
                total = entry.get('count', 0)
                avg = round(total / num_summary_months, 2) if num_summary_months > 0 else 0
                code_monthly = level_monthly.get(str(code), {})
                month_vals = [code_monthly.get(m, 0) for m in summary_months]
                row_vals = [str(code), entry.get('description', ''), total, avg] + month_vals
                for ci, v in enumerate(row_vals, 1):
                    ws_sum.cell(row=sum_row, column=ci, value=v)
                sum_row += 1

            # Grand total row for this level
            grand = sum(c.get('count', 0) for c in level_codes.values())
            grand_monthly = [
                sum(level_monthly.get(str(code), {}).get(m, 0) for code in level_codes)
                for m in summary_months
            ]
            tot_vals = ['TOTAL', '', grand,
                        round(grand / num_summary_months, 2) if num_summary_months else 0] + grand_monthly
            for ci, v in enumerate(tot_vals, 1):
                cell = ws_sum.cell(row=sum_row, column=ci, value=v)
                cell.font = sum_tot_font
                cell.fill = sum_tot_fill
            sum_row += 2  # blank separator before next level

        # Auto-fit summary sheet columns
        for col in ws_sum.columns:
            col_letter = get_column_letter(col[0].column)
            max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
            ws_sum.column_dimensions[col_letter].width = min(max_len + 3, 45)

        for level in [1, 2, 3]:
            ws = wb.create_sheet(title=level_sheet_names[level])
            level_data = data_by_level.get(level, {})
            row_num = 1

            for code in sorted(level_data.keys()):
                mfr_dict = level_data[code]
                manufacturers = sorted(mfr_dict.keys())

                # ── Code header row ──────────────────────────────────
                num_cols = 1 + len(manufacturers) + 1  # Month + mfrs + Total
                code_cell = ws.cell(row=row_num, column=1, value=f"Code: {code}")
                code_cell.font = code_font
                code_cell.fill = code_fill
                # Fill the rest of the code header row with the same background
                for c in range(2, num_cols + 1):
                    ws.cell(row=row_num, column=c).fill = code_fill
                row_num += 1

                # ── Column header row ────────────────────────────────
                headers = ['Month'] + manufacturers + ['Total']
                for col_idx, h in enumerate(headers, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = center
                row_num += 1

                # ── Monthly data rows (skip months with zero total) ──
                mfr_totals = {mfr: 0 for mfr in manufacturers}
                for month_str, month_label in zip(all_months, month_labels):
                    counts_this_month = [mfr_dict[mfr].get(month_str, 0) for mfr in manufacturers]
                    row_total = sum(counts_this_month)
                    if row_total == 0:
                        continue  # skip months with no events
                    ws.cell(row=row_num, column=1, value=month_label)
                    for col_idx, (mfr, cnt) in enumerate(zip(manufacturers, counts_this_month), 2):
                        ws.cell(row=row_num, column=col_idx, value=cnt)
                        mfr_totals[mfr] += cnt
                    ws.cell(row=row_num, column=len(headers), value=row_total)
                    row_num += 1

                # ── Total row ────────────────────────────────────────
                grand_total = sum(mfr_totals.values())
                total_vals = ['Total'] + [mfr_totals[mfr] for mfr in manufacturers] + [grand_total]
                for col_idx, v in enumerate(total_vals, 1):
                    cell = ws.cell(row=row_num, column=col_idx, value=v)
                    cell.font = total_font
                    cell.fill = total_fill
                row_num += 2  # blank separator between codes

            # ── Auto-fit column widths ───────────────────────────────
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
                ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

        # ── Patient Problem E-Codes sheet (per-manufacturer monthly) ──────
        pat_all_months = pat_mfr_monthly['months']
        pat_data_by_level = pat_mfr_monthly['data']
        pat_month_labels = [_fmt_month(m) for m in pat_all_months]

        pat_level_sheet_names = {
            1: 'Patient Problem E-Codes L1',
            2: 'Patient Problem E-Codes L2',
            3: 'Patient Problem E-Codes L3',
        }

        for level in [1, 2, 3]:
            ws_pat = wb.create_sheet(title=pat_level_sheet_names[level])
            level_data = pat_data_by_level.get(level, {})
            row_num = 1

            for code in sorted(level_data.keys()):
                mfr_dict = level_data[code]
                manufacturers = sorted(mfr_dict.keys())

                num_cols = 1 + len(manufacturers) + 1
                code_cell = ws_pat.cell(row=row_num, column=1, value=f"E-Code: {code}")
                code_cell.font = code_font
                code_cell.fill = code_fill
                for c in range(2, num_cols + 1):
                    ws_pat.cell(row=row_num, column=c).fill = code_fill
                row_num += 1

                headers = ['Month'] + manufacturers + ['Total']
                for col_idx, h in enumerate(headers, 1):
                    cell = ws_pat.cell(row=row_num, column=col_idx, value=h)
                    cell.font = hdr_font
                    cell.fill = hdr_fill
                    cell.alignment = center
                row_num += 1

                mfr_totals = {mfr: 0 for mfr in manufacturers}
                for month_str, month_label in zip(pat_all_months, pat_month_labels):
                    counts_this_month = [mfr_dict[mfr].get(month_str, 0) for mfr in manufacturers]
                    row_total = sum(counts_this_month)
                    if row_total == 0:
                        continue
                    ws_pat.cell(row=row_num, column=1, value=month_label)
                    for col_idx, (mfr, cnt) in enumerate(zip(manufacturers, counts_this_month), 2):
                        ws_pat.cell(row=row_num, column=col_idx, value=cnt)
                        mfr_totals[mfr] += cnt
                    ws_pat.cell(row=row_num, column=len(headers), value=row_total)
                    row_num += 1

                grand_total = sum(mfr_totals.values())
                total_vals = ['Total'] + [mfr_totals[mfr] for mfr in manufacturers] + [grand_total]
                for col_idx, v in enumerate(total_vals, 1):
                    cell = ws_pat.cell(row=row_num, column=col_idx, value=v)
                    cell.font = total_font
                    cell.fill = total_fill
                row_num += 2

            for col in ws_pat.columns:
                col_letter = get_column_letter(col[0].column)
                max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
                ws_pat.column_dimensions[col_letter].width = min(max_len + 3, 45)

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name='imdrf_code_counts_all_levels.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/imdrf-insights/generate-pdf-report', methods=['POST'])
@login_required
def imdrf_insights_generate_pdf():
    """Generate a PDF trend analysis report for a selected manufacturer and IMDRF code."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid request body'}), 400

    file_id      = data.get('file_id', '').strip()
    hist_file_id = data.get('hist_file_id', '').strip()
    manufacturer = data.get('manufacturer', '').strip()
    code_filter  = (data.get('code', 'ALL') or 'ALL').strip()
    period_from  = data.get('period_from', '')   # "YYYY-MM-DD"
    period_to    = data.get('period_to', '')
    grain        = data.get('grain', 'M')
    level        = int(data.get('level', 1))

    if not file_id or not hist_file_id:
        return jsonify({'error': 'Missing file_id or hist_file_id'}), 400
    if not manufacturer:
        return jsonify({'error': 'Manufacturer is required'}), 400
    if not period_from or not period_to:
        return jsonify({'error': 'Period dates are required'}), 400

    # Load file paths from session
    main_meta = session.get(f'insights_file_{file_id}')
    hist_meta = session.get(f'insights_file_{hist_file_id}')

    if not main_meta:
        return jsonify({'error': 'Current period data not found. Please reload your data.'}), 404
    if not hist_meta:
        return jsonify({'error': 'Historical data not found. Please try generating the report again.'}), 404

    main_path = main_meta.get('path') if isinstance(main_meta, dict) else main_meta
    hist_path = hist_meta.get('path') if isinstance(hist_meta, dict) else hist_meta

    if not main_path or not os.path.exists(main_path):
        return jsonify({'error': 'Current period file not found. Please reload your data.'}), 404
    if not hist_path or not os.path.exists(hist_path):
        return jsonify({'error': 'Historical data file not found. Please try again.'}), 404

    try:
        from backend.imdrf_insights import (
            prepare_data_for_insights,
            compute_report_data,
            render_trend_chart,
            build_report_pdf,
        )

        main_result = prepare_data_for_insights(main_path, level=level)
        hist_result = prepare_data_for_insights(hist_path, level=level)

        df_current = main_result['df_exploded']
        df_hist    = hist_result['df_exploded']
        mfr_col    = main_result.get('mfr_col') or '_manufacturer'

        # Ensure manufacturer column exists
        if mfr_col not in df_current.columns:
            df_current['_manufacturer'] = 'All Data'
            mfr_col = '_manufacturer'
        if mfr_col not in df_hist.columns:
            df_hist[mfr_col] = 'All Data'

        report_data = compute_report_data(
            df_current, df_hist, mfr_col, manufacturer,
            code_filter, period_from, period_to, grain, level
        )

        if not report_data['trends']:
            return jsonify({'error': 'No data found for the selected manufacturer and code in the specified period.'}), 400

        # Generate chart images
        chart_images = []
        for entry in report_data['trends'].values():
            buf = render_trend_chart(
                entry['code'], entry['labels'],
                entry['mfr_values'], entry['peers_values'],
                manufacturer, report_data['grain_label']
            )
            chart_images.append(buf)

        pdf_bytes = build_report_pdf(report_data, chart_images)

        safe_mfr = re.sub(r'[^A-Za-z0-9_-]', '_', manufacturer)[:30]
        filename  = f"IMDRF_Report_{safe_mfr}_{period_from}_{period_to}.pdf"

        return send_file(
            io.BytesIO(pdf_bytes),
            mimetype='application/pdf',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/imdrf-insights/compute-proportions', methods=['POST'])
@login_required
def imdrf_insights_compute_proportions():
    """Compute proportion of a specific IMDRF code across combined historical + current period."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid request body'}), 400

    file_id      = data.get('file_id', '').strip()
    hist_file_id = data.get('hist_file_id', '').strip()
    code         = data.get('code', '').strip()
    period_from  = data.get('period_from', '')
    period_to    = data.get('period_to', '')
    level        = int(data.get('level', 1))

    if not file_id or not hist_file_id:
        return jsonify({'error': 'Missing file_id or hist_file_id'}), 400
    if not code or code.upper() == 'ALL':
        return jsonify({'error': 'A specific event code is required (not ALL)'}), 400
    if not period_from or not period_to:
        return jsonify({'error': 'Period dates are required'}), 400

    main_meta = session.get(f'insights_file_{file_id}')
    hist_meta = session.get(f'insights_file_{hist_file_id}')
    if not main_meta:
        return jsonify({'error': 'Current period data not found. Please reload your data.'}), 404
    if not hist_meta:
        return jsonify({'error': 'Historical data not found. Please generate the report first.'}), 404

    main_path = main_meta.get('path') if isinstance(main_meta, dict) else main_meta
    hist_path = hist_meta.get('path') if isinstance(hist_meta, dict) else hist_meta

    if not main_path or not os.path.exists(main_path):
        return jsonify({'error': 'Current period file not found. Please reload your data.'}), 404
    if not hist_path or not os.path.exists(hist_path):
        return jsonify({'error': 'Historical data file not found. Please try again.'}), 404

    try:
        from backend.imdrf_insights import prepare_data_for_insights, compute_proportions

        main_result = prepare_data_for_insights(main_path, level=level)
        hist_result = prepare_data_for_insights(hist_path, level=level)

        result = compute_proportions(
            main_result['df_exploded'],
            hist_result['df_exploded'],
            code, period_from, period_to
        )
        return jsonify(result)

    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/imdrf-insights/hist-code-table', methods=['POST'])
@login_required
def imdrf_insights_hist_code_table():
    """Return IMDRF code distribution table for the historical (2-year) dataset."""
    data = request.get_json()
    if not data:
        return jsonify({'error': 'Invalid request body'}), 400

    hist_file_id = data.get('hist_file_id', '').strip()
    level        = int(data.get('level', 1))

    if not hist_file_id:
        return jsonify({'error': 'Missing hist_file_id'}), 400

    hist_meta = session.get(f'insights_file_{hist_file_id}')
    if not hist_meta:
        return jsonify({'error': 'Historical data not found in session. Please reload.'}), 404

    hist_path = hist_meta.get('path') if isinstance(hist_meta, dict) else hist_meta
    if not hist_path or not os.path.exists(hist_path):
        return jsonify({'error': 'Historical data file not found on disk.'}), 404

    try:
        from backend.imdrf_insights import prepare_data_for_insights, get_hist_code_table
        hist_result = prepare_data_for_insights(hist_path, level=level)
        result = get_hist_code_table(hist_result['df_exploded'])
        return jsonify(result)
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# MAUDE Bulk Export (openFDA)
def _format_openfda_search_value(value: str) -> str:
    """Format a value for openFDA search queries, quoting if needed."""
    if re.fullmatch(r'[A-Za-z0-9]+', value):
        return value
    escaped = value.replace('"', '\\"')
    return f'"{escaped}"'


def _parse_next_link(link_header: str) -> str:
    if not link_header:
        return None
    parts = link_header.split(',')
    for part in parts:
        if re.search(r'rel\s*=\s*"?next"?', part, re.IGNORECASE):
            match = re.search(r'<([^>]+)>', part)
            if match:
                return match.group(1)
    return None


def _build_product_code_query(field_or_fields, code_value: str) -> str:
    if isinstance(field_or_fields, (list, tuple)):
        clauses = [f"{field}:{code_value}" for field in field_or_fields]
        return f"({ ' OR '.join(clauses) })"
    return f"{field_or_fields}:{code_value}"


def _parse_openfda_date(value):
    if not value:
        return None
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y%m%d', '%Y-%m-%d', '%Y/%m/%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _format_openfda_date(value):
    parsed = _parse_openfda_date(value)
    return parsed.isoformat() if parsed else (value or '')


def _normalize_code_value(value):
    if value is None:
        return []
    if isinstance(value, list):
        values = value
    else:
        values = [value]
    codes = []
    for item in values:
        text = _stringify_value(item).upper().strip()
        if not text:
            continue
        codes.extend([c for c in re.split(r'[^A-Z0-9]+', text) if c])
    return _unique_preserve_order(codes)


def _record_has_product_code(record, code: str) -> bool:
    code = (code or '').upper().strip()
    if not code:
        return False
    devices = record.get('device') or []
    candidates = []
    for device in devices:
        candidates.extend(_normalize_code_value(device.get('device_report_product_code')))
        candidates.extend(_normalize_code_value(device.get('product_code')))
        openfda = device.get('openfda') or {}
        candidates.extend(_normalize_code_value(openfda.get('product_code')))
    openfda_root = record.get('openfda') or {}
    candidates.extend(_normalize_code_value(openfda_root.get('product_code')))
    return code in set(candidates)


def _record_in_date_range(record, start_date, end_date, date_field: str = None) -> bool:
    value = record.get(date_field) if date_field else None
    if not value:
        value = record.get('date_received')
    parsed = _parse_openfda_date(value)
    if not parsed:
        return False
    return start_date <= parsed <= end_date


def _get_mdr_text(record, text_type_code: str) -> str:
    for item in record.get('mdr_text') or []:
        if item.get('text_type_code') == text_type_code:
            return item.get('text', '')
    return ''


def _ensure_api_key(url: str, api_key: str) -> str:
    if not api_key:
        return url
    parsed = urlparse(url)
    query = parse_qs(parsed.query)
    if 'api_key' not in query:
        query['api_key'] = [api_key]
        new_query = urlencode(query, doseq=True)
        parsed = parsed._replace(query=new_query)
        return urlunparse(parsed)
    return url


_OPENFDA_HEADERS = {
    'User-Agent': 'Mozilla/5.0 (compatible; FDA-MAUDE-Analysis/1.0; +https://open.fda.gov)',
    'Accept': 'application/json',
}


def _openfda_get(url: str, params: dict = None, max_retries: int = 4, timeout=(5, 30), allow_not_found: bool = False, allow_bad_request: bool = False):
    """GET with retries on 429/5xx and exponential backoff.

    If OPENFDA_API_KEY is set, injects it automatically. On API_KEY_INVALID (key not yet
    working), transparently retries without the key so anonymous calls still succeed.
    """
    api_key = os.getenv('OPENFDA_API_KEY', '').strip()
    # Inject key if available and not already present
    if api_key and params and 'api_key' not in params:
        params = {**params, 'api_key': api_key}

    for attempt in range(max_retries + 1):
        try:
            response = requests.get(url, params=params, headers=_OPENFDA_HEADERS, timeout=timeout)
            if response.status_code == 404 and allow_not_found:
                return response
            if response.status_code == 400 and allow_bad_request:
                return response
            if response.status_code == 403:
                # If the key is invalid, retry once without it so anonymous still works
                try:
                    err_code = response.json().get('error', {}).get('code', '')
                except Exception:
                    err_code = ''
                if err_code == 'API_KEY_INVALID' and params and 'api_key' in params:
                    anon_params = {k: v for k, v in params.items() if k != 'api_key'}
                    return requests.get(url, params=anon_params, headers=_OPENFDA_HEADERS, timeout=timeout)
                return response
            if response.status_code in (429, 500, 502, 503, 504):
                if attempt < max_retries:
                    retry_after = response.headers.get('Retry-After')
                    delay = float(retry_after) if retry_after and retry_after.isdigit() else (2 ** attempt)
                    time.sleep(delay)
                    continue
                return response
            response.raise_for_status()
            return response
        except requests.RequestException:
            if attempt < max_retries:
                time.sleep(2 ** attempt)
                continue
            raise




def _normalize_to_list(value):
    if value is None:
        return []
    if isinstance(value, list):
        return value
    return [value]


def _stringify_value(value):
    if value is None:
        return ''
    if isinstance(value, dict):
        for key in ('device_problem_code', 'device_problem_text', 'code', 'text', 'value', 'description'):
            if key in value:
                return str(value.get(key) or '')
        if len(value) == 1:
            return str(next(iter(value.values())))
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _unique_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        if value in seen or value == '':
            continue
        seen.add(value)
        result.append(value)
    return result


def _collect_device_field(devices, field):
    values = []
    for device in devices:
        for value in _normalize_to_list(device.get(field)):
            values.append(_stringify_value(value))
    return _unique_preserve_order(values)


def _collect_device_openfda_field(devices, field):
    values = []
    for device in devices:
        openfda = device.get('openfda') or {}
        for value in _normalize_to_list(openfda.get(field)):
            values.append(_stringify_value(value))
    return _unique_preserve_order(values)


def _maude_probe(product_code, date_from_str, date_to_str):
    """Probe openFDA to find the correct search field/code/date combination.

    Returns (error_dict_or_None, probe_dict_or_None).
    error_dict: {'error': str, 'status_code': int}
    probe_dict: all variables needed by _maude_generate_csv and api_maude_export.
    """
    try:
        start_date = datetime.strptime(date_from_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(date_to_str, '%Y-%m-%d').date()
    except ValueError:
        return {'error': 'Invalid date format. Use YYYY-MM-DD.', 'status_code': 400}, None

    if start_date > end_date:
        return {'error': 'Date From must be on or before Date To.', 'status_code': 400}, None

    start_openfda = start_date.strftime('%Y%m%d')
    end_openfda = end_date.strftime('%Y%m%d')
    normalized_code = product_code.upper()
    candidates = [normalized_code]
    if '-' in normalized_code:
        candidates.append(normalized_code.replace('-', ''))
    if ' ' in normalized_code:
        candidates.append(normalized_code.replace(' ', ''))
    candidates = [c for i, c in enumerate(candidates) if c and c not in candidates[:i]]

    api_key = os.getenv('OPENFDA_API_KEY')
    base_url = 'https://api.fda.gov/device/event.json'
    limit = 1000
    sort_field = 'date_received'
    sort = f'{sort_field}:asc'
    search_fields = [
        'device.device_report_product_code',
        'device.product_code',
        'device.openfda.product_code',
        'openfda.product_code'
    ]
    date_fields = ['date_received', 'date_report', 'date_of_event']
    selected_field = None
    selected_code = None
    selected_date_field = None
    rate_limited = False
    rate_limit_status = 429  # will hold 403 or 429

    # Primary probe: device.device_report_product_code + date_received
    primary_search = (
        f'date_received:[{start_openfda} TO {end_openfda}] AND '
        f'device.device_report_product_code:{_format_openfda_search_value(normalized_code)}'
    )
    def _probe_one(params):
        """Single probe call — fail fast, no retry."""
        nonlocal rate_limited, rate_limit_status
        resp = _openfda_get(base_url, params=params, max_retries=0, allow_not_found=True, allow_bad_request=True)
        if resp.status_code in (403, 429):
            rate_limited = True
            rate_limit_status = resp.status_code
        return resp

    primary_params = {'search': primary_search, 'limit': 1, 'sort': sort}
    try:
        primary_resp = _probe_one(primary_params)
        if primary_resp.status_code == 200:
            primary_data = primary_resp.json()
            if primary_data.get('results'):
                selected_field = 'device.device_report_product_code'
                selected_code = normalized_code
                selected_date_field = 'date_received'
    except Exception:
        pass

    if not selected_field and not rate_limited:
        for date_field in date_fields:
            if rate_limited:
                break
            for code_candidate in candidates:
                if rate_limited:
                    break
                search_value = _format_openfda_search_value(code_candidate)
                for field in search_fields:
                    search = f'{date_field}:[{start_openfda} TO {end_openfda}] AND {field}:{search_value}'
                    try:
                        response = _probe_one({'search': search, 'limit': 1, 'sort': sort})
                        if rate_limited:
                            break
                        if response.status_code in (400, 404):
                            continue
                        data_probe = response.json()
                        if data_probe.get('results'):
                            selected_field = field
                            selected_code = code_candidate
                            selected_date_field = date_field
                            break
                    except Exception:
                        continue
                if selected_field or rate_limited:
                    break
            if selected_field or rate_limited:
                break

    combined_fields = search_fields
    if not selected_field and not rate_limited:
        for date_field in date_fields:
            if rate_limited:
                break
            for code_candidate in candidates:
                search_value = _format_openfda_search_value(code_candidate)
                combined_query = _build_product_code_query(combined_fields, search_value)
                search = f'{date_field}:[{start_openfda} TO {end_openfda}] AND {combined_query}'
                try:
                    response = _probe_one({'search': search, 'limit': 1, 'sort': sort})
                    if rate_limited:
                        break
                    if response.status_code in (400, 404):
                        continue
                    data_probe = response.json()
                    if data_probe.get('results'):
                        selected_field = combined_fields
                        selected_code = code_candidate
                        selected_date_field = date_field
                        break
                except Exception:
                    continue
            if selected_field or rate_limited:
                break

    if not selected_field and rate_limited:
        if rate_limit_status == 403:
            msg = 'openFDA returned 403 Forbidden. The API key may not be activated yet or the IP is blocked. Please verify the API key at open.fda.gov.'
        else:
            msg = 'openFDA is rate-limiting this server (429). Please wait a minute and try again.'
        return {'error': msg, 'status_code': rate_limit_status}, None

    if not selected_field:
        return {
            'error': (
                f'No MAUDE events found for {product_code} between {date_from_str} and {date_to_str}. '
                'Please confirm the product code and date range, then try again.'
            ),
            'status_code': 404
        }, None

    date_filtered_search = True

    if not selected_code:
        selected_code = normalized_code
    if not selected_date_field:
        selected_date_field = 'date_received'
    sort_field = selected_date_field
    sort = f'{sort_field}:asc'
    search_value = _format_openfda_search_value(selected_code)
    product_query = _build_product_code_query(selected_field, search_value)
    if date_filtered_search:
        search = f'{selected_date_field}:[{start_openfda} TO {end_openfda}] AND {product_query}'
    else:
        search = product_query
    fetch_params = {'search': search, 'limit': limit, 'sort': sort}

    try:
        first_response = _openfda_get(base_url, params=fetch_params, allow_not_found=True, allow_bad_request=True)
        if first_response.status_code == 400:
            # Retry without sort in case sort field is unsupported
            first_response = _openfda_get(base_url, params={'search': search, 'limit': limit}, allow_not_found=True, allow_bad_request=True)
        if first_response.status_code == 403:
            try:
                err_code = first_response.json().get('error', {}).get('code', '')
            except Exception:
                err_code = ''
            if err_code == 'API_KEY_MISSING':
                msg = ('openFDA requires an API key for this query (too many results without a date filter). '
                       'Register a free key at open.fda.gov/apis/authentication/ and add it to your .env as OPENFDA_API_KEY.')
            else:
                msg = 'openFDA returned 403 Forbidden. The API key may not be activated yet or the IP is blocked. Please verify the API key at open.fda.gov.'
            return {'error': msg, 'status_code': 403}, None
        if first_response.status_code == 429:
            return {
                'error': 'openFDA is rate-limiting this server (429). Please wait a minute and try again.',
                'status_code': 429
            }, None
        if first_response.status_code == 400:
            return {
                'error': 'openFDA rejected the query. Please verify the product code and date range and try again.',
                'status_code': 400
            }, None
        if first_response.status_code == 404:
            return {
                'error': 'No matching records found for the provided product code and date range.',
                'status_code': 404
            }, None
        first_payload = first_response.json()
    except Exception as e:
        return {'error': f'openFDA request failed: {str(e)}', 'status_code': 502}, None

    first_results = first_payload.get('results', [])
    total_results = None
    try:
        total_results = first_payload.get('meta', {}).get('results', {}).get('total')
    except Exception:
        total_results = None
    if not first_results:
        return {
            'error': 'No matching records found for the provided product code and date range.',
            'status_code': 404
        }, None

    next_url = _parse_next_link(first_response.headers.get('Link'))

    if not date_filtered_search:
        total_results = None

    columns = [
        'Unnamed: 0', 'Web Address', 'Report Number', 'Event Date', 'Event Type',
        'Manufacturer', 'Date Received', 'Product Code', ' Brand Name',
        ' Device Problem', 'Patient Problem', 'PMA/PMN Number', 'Exemption Number',
        'Number of Events', 'Event Text'
    ]

    return None, {
        'first_results': first_results,
        'next_url': next_url,
        'search': search,
        'sort': sort,
        'limit': limit,
        'start_date': start_date,
        'end_date': end_date,
        'selected_date_field': selected_date_field,
        'selected_code': selected_code,
        'product_query': product_query,  # stored for year-by-year re-fetching
        'base_url': base_url,
        'columns': columns,
        'total_results': total_results,
        'product_code': product_code,
        'date_from': date_from_str,
        'date_to': date_to_str,
    }


def _make_year_probe(base_probe, year_start_str, year_end_str):
    """Build a year-scoped probe reusing already-detected fields from base_probe.

    Returns a probe dict ready for _maude_generate_csv, or None if no data for that year.
    """
    try:
        start_date = datetime.strptime(year_start_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(year_end_str, '%Y-%m-%d').date()
    except ValueError:
        return None

    start_openfda = start_date.strftime('%Y%m%d')
    end_openfda = end_date.strftime('%Y%m%d')

    selected_date_field = base_probe['selected_date_field']
    product_query = base_probe['product_query']
    base_url = base_probe['base_url']
    limit = base_probe['limit']
    sort = base_probe['sort']

    search = f'{selected_date_field}:[{start_openfda} TO {end_openfda}] AND {product_query}'
    fetch_params = {'search': search, 'limit': limit, 'sort': sort}

    try:
        response = _openfda_get(base_url, params=fetch_params, allow_not_found=True, allow_bad_request=True)
        if response.status_code in (404, 400):
            return None
        payload = response.json()
        first_results = payload.get('results', [])
        if not first_results:
            return None
        total = payload.get('meta', {}).get('results', {}).get('total')
        next_url = _parse_next_link(response.headers.get('Link'))
    except Exception:
        return None

    return {
        'first_results': first_results,
        'next_url': next_url,
        'search': search,
        'sort': sort,
        'limit': limit,
        'start_date': start_date,
        'end_date': end_date,
        'selected_date_field': selected_date_field,
        'selected_code': base_probe['selected_code'],
        'product_query': product_query,
        'base_url': base_url,
        'columns': base_probe['columns'],
        'total_results': total,
    }


def _maude_build_row(record, selected_code):
    """Build a CSV row dict from a single openFDA device event record."""
    devices = record.get('device') or []

    report_product_codes = _collect_device_field(devices, 'device_report_product_code')
    brand_names = _unique_preserve_order(
        _collect_device_field(devices, 'brand_name') + _collect_device_openfda_field(devices, 'brand_name')
    )
    pma_numbers = _collect_device_openfda_field(devices, 'pma_number')
    k_numbers = _collect_device_openfda_field(devices, 'k_number')
    exemption_numbers = _collect_device_openfda_field(devices, 'exemption_number')

    device_problem_values = []
    for value in _normalize_to_list(record.get('product_problems')):
        device_problem_values.append(_stringify_value(value))
    for value in _normalize_to_list(record.get('device_problem')):
        device_problem_values.append(_stringify_value(value))
    for device in devices:
        for value in _normalize_to_list(device.get('device_problem')):
            device_problem_values.append(_stringify_value(value))
    device_problem_values = _unique_preserve_order([v.strip() for v in device_problem_values if v.strip()])

    patient_problem_values = []
    for value in _normalize_to_list(record.get('patient_problem_text')):
        patient_problem_values.append(_stringify_value(value))
    for patient in record.get('patient') or []:
        for value in _normalize_to_list(patient.get('patient_problems')):
            patient_problem_values.append(_stringify_value(value))
    patient_problem_values = _unique_preserve_order([v.strip() for v in patient_problem_values if v.strip()])

    report_number = record.get('report_number') or record.get('mdr_report_key') or ''
    manufacturer_name = record.get('manufacturer_name', '')
    if not manufacturer_name:
        manufacturer_candidates = _collect_device_field(devices, 'manufacturer_d_name')
        if manufacturer_candidates:
            manufacturer_name = '; '.join(manufacturer_candidates)
    mdr_key = record.get('mdr_report_key') or ''
    event_date = (
        record.get('date_of_event')
        or record.get('date_report')
        or record.get('date_report_to_fda')
        or record.get('date_received')
    )
    event_description = record.get('event_description', '') or _get_mdr_text(record, 'Description of Event or Problem')
    if event_description and not str(event_description).strip().lower().startswith('event description:'):
        event_description = f"Event Description: {event_description}"
    product_code_value = selected_code or (report_product_codes[0] if report_product_codes else '')
    pma_pmn_numbers = _unique_preserve_order(pma_numbers + k_numbers)

    web_address = ''
    if mdr_key:
        web_address = (
            f"https://www.accessdata.fda.gov/scripts/cdrh/cfdocs/cfMAUDE/Detail.CFM"
            f"?MDRFOI__ID={mdr_key}&pc={product_code_value}"
        )

    return {
        'Unnamed: 0': '',
        'Web Address': web_address,
        'Report Number': report_number,
        'Event Date': _format_openfda_date(event_date),
        'Event Type': record.get('event_type', ''),
        'Manufacturer': manufacturer_name,
        'Date Received': _format_openfda_date(record.get('date_received')),
        'Product Code': product_code_value,
        ' Brand Name': '; '.join(brand_names),
        ' Device Problem': '; '.join(device_problem_values),
        'Patient Problem': '; '.join(patient_problem_values),
        'PMA/PMN Number': '; '.join(pma_pmn_numbers),
        'Exemption Number': '; '.join(exemption_numbers),
        'Number of Events': record.get('number_devices_in_event') or record.get('number_patients_in_event') or '1',
        'Event Text': event_description
    }


def _maude_generate_csv(probe, progress_callback=None):
    """Generator that yields CSV bytes for a MAUDE export given a completed probe dict."""
    columns = probe['columns']
    first_results = probe['first_results']
    next_url = probe['next_url']
    search = probe['search']
    sort = probe['sort']
    limit = probe['limit']
    start_date = probe['start_date']
    end_date = probe['end_date']
    selected_date_field = probe['selected_date_field']
    selected_code = probe['selected_code']
    base_url = probe['base_url']

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns)
    writer.writeheader()
    yield output.getvalue().encode('utf-8')
    output.seek(0)
    output.truncate(0)

    current_results = first_results
    current_next = next_url
    last_next = None
    processed = 0
    scanned = 0
    pages_fetched = 1
    last_progress_time = time.time()
    buffered_rows = 0
    flush_every = 200
    max_skip = 25000

    while True:
        for record in current_results:
            scanned += 1
            if not _record_in_date_range(record, start_date, end_date, selected_date_field):
                if progress_callback and scanned % 100 == 0:
                    progress_callback(processed, scanned)
                continue
            if not _record_has_product_code(record, selected_code):
                if progress_callback and scanned % 100 == 0:
                    progress_callback(processed, scanned)
                continue
            writer.writerow(_maude_build_row(record, selected_code))
            buffered_rows += 1
            processed += 1
            if buffered_rows >= flush_every:
                yield output.getvalue().encode('utf-8')
                output.seek(0)
                output.truncate(0)
                buffered_rows = 0
            if progress_callback and (processed % 20 == 0 or (time.time() - last_progress_time) > 0.5):
                progress_callback(processed, scanned)
                last_progress_time = time.time()

        if buffered_rows > 0:
            yield output.getvalue().encode('utf-8')
            output.seek(0)
            output.truncate(0)
            buffered_rows = 0

        if not current_next:
            if len(current_results) == limit:
                skip_val = pages_fetched * limit
                if skip_val <= max_skip:
                    skip_params = {
                        'search': search,
                        'limit': limit,
                        'sort': sort,
                        'skip': skip_val,
                    }
                    current_next = requests.Request('GET', base_url, params=skip_params).prepare().url
                else:
                    break
            else:
                break

        if current_next == last_next:
            break

        try:
            response = _openfda_get(current_next, params=None)
            payload = response.json()
        except Exception:
            break

        current_results = payload.get('results', [])
        if not current_results:
            break

        pages_fetched += 1
        last_next = current_next
        current_next = _parse_next_link(response.headers.get('Link'))


@app.route('/api/maude/export', methods=['POST'])
@login_required
def api_maude_export():
    data = request.get_json() or {}
    product_code = (data.get('product_code') or '').strip().upper()
    date_from = (data.get('date_from') or '').strip()
    date_to = (data.get('date_to') or '').strip()
    fmt = (data.get('format') or 'csv').strip().lower()
    async_requested = bool(data.get('async'))

    if not product_code:
        return jsonify({'error': 'Product code is required.'}), 400

    if fmt != 'csv':
        return jsonify({'error': 'Only CSV format is supported at this time.'}), 400

    err, probe = _maude_probe(product_code, date_from, date_to)
    if err:
        return jsonify({'error': err['error']}), err['status_code']

    safe_code = re.sub(r'[^A-Za-z0-9_-]+', '', product_code.upper()) or 'CODE'
    filename = f"maude_{safe_code}_{date_from}_{date_to}.csv"

    if async_requested:
        job_id = str(uuid.uuid4())
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"export_{job_id}.csv")
        user_id = current_user.id
        total_results = probe['total_results']
        _set_export_status(job_id, status='queued', user_id=user_id, output_filename=filename,
                           processed=0, scanned=0, total=total_results)

        def run_export_job(job_user_id):
            _set_export_status(job_id, status='running', user_id=job_user_id)
            def progress_callback(count, scanned_count):
                _set_export_status(job_id, processed=count, scanned=scanned_count,
                                   total=total_results, user_id=job_user_id)
            try:
                with open(output_path, 'wb') as f:
                    for chunk in _maude_generate_csv(probe, progress_callback=progress_callback):
                        f.write(chunk)
                _set_export_status(job_id, status='done', output_path=output_path,
                                   output_filename=filename, user_id=job_user_id)
            except Exception as e:
                _set_export_status(job_id, status='failed', error=str(e), user_id=job_user_id)

        thread = threading.Thread(target=run_export_job, args=(user_id,), daemon=True)
        thread.start()
        return jsonify({
            'job_id': job_id,
            'status_url': url_for('maude_export_status', job_id=job_id),
            'download_url': url_for('maude_export_download', job_id=job_id)
        }), 202

    response = Response(stream_with_context(_maude_generate_csv(probe)), mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _sanitize_json(obj):
    """Recursively replace NaN/Inf float values with None so jsonify produces valid JSON."""
    import math
    if isinstance(obj, dict):
        return {k: _sanitize_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_json(v) for v in obj]
    if isinstance(obj, float) and (math.isnan(obj) or math.isinf(obj)):
        return None
    return obj


def _fmt_num(n):
    """Format a number with thousands-separator commas, or 'N/A' for None."""
    if n is None:
        return 'N/A'
    try:
        return f"{int(n):,}"
    except (TypeError, ValueError):
        return str(n)


def _build_audit_report(audit):
    """Build a plain-text audit report string from stored pipeline audit_data."""
    W = 79
    SEP  = '─' * W
    THICK = '═' * W
    lines = []

    lines.append(THICK)
    lines.append('     FDA MAUDE DATA PROCESSING  —  AUDIT REPORT'.center(W))
    lines.append(THICK)
    lines.append('')

    run_at = audit.get('run_at', '')
    if run_at:
        try:
            dt = datetime.fromisoformat(run_at)
            run_at = dt.strftime('%d-%m-%Y  %H:%M:%S  UTC')
        except Exception:
            pass

    pipeline_type = audit.get('pipeline_type', 'N/A')
    type_label = {
        'raw':   'Raw  (Download only)',
        'clean': 'Clean  (Download + Process)',
        'full':  'Full  (Download + Process + IMDRF Code Counts)',
    }.get(pipeline_type, pipeline_type)

    lines.append(f"  Report Generated  : {run_at}")
    lines.append(f"  Product Code      : {audit.get('product_code', 'N/A')}")
    lines.append(f"  Date Range        : {audit.get('date_from', 'N/A')}  to  {audit.get('date_to', 'N/A')}")
    lines.append(f"  Pipeline Type     : {type_label}")
    lines.append('')

    # ── Section 1: Data Source ──────────────────────────────────────────────
    lines.append(SEP)
    lines.append('  SECTION 1  —  DATA SOURCE')
    lines.append(SEP)
    lines.append(f"  Total Records Found (openFDA)  : {_fmt_num(audit.get('total_records_found'))}")
    lines.append('')

    # ── Section 2: Raw Download ─────────────────────────────────────────────
    lines.append(SEP)
    lines.append('  SECTION 2  —  RAW DATA DOWNLOAD')
    lines.append(SEP)
    raw_dl = audit.get('raw_records_downloaded') or audit.get('total_records_found')
    lines.append(f"  Records Downloaded  : {_fmt_num(raw_dl)}")
    lines.append('')

    stats = audit.get('stats')
    if not stats:
        # raw pipeline — no cleaning section
        lines.append(THICK)
        lines.append('                    END OF AUDIT REPORT'.center(W))
        lines.append(THICK)
        return '\n'.join(lines)

    # ── Section 3: Data Cleaning ────────────────────────────────────────────
    lines.append(SEP)
    lines.append('  SECTION 3  —  DATA CLEANING')
    lines.append(SEP)
    orig    = stats.get('original_rows', 0)
    final   = stats.get('final_rows', 0)
    removed = stats.get('rows_removed', 0)
    lines.append(f"  Original Rows    : {_fmt_num(orig)}")
    lines.append(f"  Final Rows       : {_fmt_num(final)}")
    lines.append(f"  Rows Removed     : {_fmt_num(removed)}")
    lines.append('')

    rows_by_reason = stats.get('rows_removed_by_reason', [])
    lines.append('  Row Removal Breakdown:')
    if rows_by_reason:
        for entry in rows_by_reason:
            lines.append(f"    •  {entry['reason']}  :  {_fmt_num(entry['count'])} rows")
    else:
        lines.append('    •  None')
    lines.append('')

    orig_cols  = stats.get('original_cols', 0)
    final_cols = stats.get('final_cols', 0)
    cols_removed = stats.get('cols_removed', [])
    lines.append(f"  Columns Before   : {orig_cols}")
    lines.append(f"  Columns After    : {final_cols}")
    if cols_removed:
        # wrap long list at ~70 chars
        col_str = ', '.join(cols_removed)
        lines.append(f"  Columns Removed ({len(cols_removed)})  :  {col_str}")
    else:
        lines.append('  Columns Removed  : None')
    lines.append('  Column Added     : IMDRF Code  (inserted adjacent to Device Problem)')
    lines.append('')

    # ── Section 4: IMDRF Mapping ────────────────────────────────────────────
    imdrf_stats = stats.get('imdrf_stats', {})
    if imdrf_stats:
        lines.append(SEP)
        lines.append('  SECTION 4  —  IMDRF CODE MAPPING')
        lines.append(SEP)
        lines.append(f"  Non-empty Device Problems  : {_fmt_num(imdrf_stats.get('non_empty_device_problems', 0))}")
        lines.append(f"  IMDRF Codes Mapped         : {_fmt_num(imdrf_stats.get('mapped', 0))}")
        lines.append(f"  Unmapped (left blank)      : {_fmt_num(imdrf_stats.get('unmapped', 0))}")
        lines.append('')

    # ── Section 5: A24/A25 Exclusion ───────────────────────────────────────
    a24_a25_desc  = audit.get('a24_a25_excluded', {})
    a24_a25_count = audit.get('a24_a25_row_count', 0)
    if a24_a25_desc:
        lines.append(SEP)
        lines.append('  SECTION 5  —  EXCLUDED IMDRF CODES  (A24 / A25 FILTER)')
        lines.append(SEP)
        lines.append(
            f"  Rows with A24/A25 IMDRF codes excluded from Code Counts output  :  "
            f"{_fmt_num(a24_a25_count)}"
        )
        lines.append('')
        lines.append('  Excluded Code Descriptions  (from IMDRF Annexes A-G):')
        for code, desc in sorted(a24_a25_desc.items()):
            lines.append(f"    •  {code}  —  {desc}")
        lines.append('  Reason  : These codes are internally excluded per project configuration.')
        lines.append('')

    # ── Section 6: Per-Level Code Count Summary ─────────────────────────────
    level_summary = audit.get('level_summary', {})
    if level_summary:
        lines.append(SEP)
        lines.append('  SECTION 6  —  IMDRF CODE COUNT SUMMARY  (PER LEVEL)')
        lines.append(SEP)
        for lvl in ['1', '2', '3']:
            lvl_data  = level_summary.get(lvl, {})
            distinct  = lvl_data.get('distinct_codes', 0)
            total_occ = lvl_data.get('total_instances', 0)
            lines.append(
                f"  Level {lvl}  :  {_fmt_num(distinct):>6} distinct codes"
                f"   |   {_fmt_num(total_occ):>8} total occurrences"
            )
        lines.append('')

    # ── Section 7: Manufacturer List ────────────────────────────────────────
    mfr_list = stats.get('manufacturer_list', [])
    if mfr_list:
        lines.append(SEP)
        lines.append('  SECTION 7  —  MANUFACTURER LIST')
        lines.append(SEP)
        lines.append(f"  Total Unique Manufacturers  : {len(mfr_list)}")
        lines.append('')
        for i, mfr in enumerate(mfr_list[:500], 1):
            lines.append(f"  {i:>4}.  {mfr}")
        if len(mfr_list) > 500:
            lines.append(f"         ... and {len(mfr_list) - 500} more (truncated for brevity)")
        lines.append('')

    # ── Section 8: Validation Results ──────────────────────────────────────
    validation = stats.get('validation', {})
    if validation:
        lines.append(SEP)
        lines.append('  SECTION 8  —  VALIDATION RESULTS')
        lines.append(SEP)
        checks = [
            ('Column Count Check',    'column_count_correct'),
            ('No Timestamps Check',   'no_timestamps'),
            ('Date Format Check',     'date_format_correct'),
            ('IMDRF Adjacency Check', 'imdrf_adjacent'),
            ('IMDRF Codes Valid',     'imdrf_codes_valid'),
            ('File Integrity Check',  'file_will_open'),
        ]
        all_passed = True
        for label, key in checks:
            val = validation.get(key)
            if val is None:
                status_str = 'N/A'
            elif val:
                status_str = 'PASSED'
            else:
                status_str = 'FAILED'
                all_passed = False
            lines.append(f"  {label:<28} : {status_str}")
        lines.append('')
        overall = 'ALL CHECKS PASSED' if all_passed else 'ONE OR MORE CHECKS FAILED'
        lines.append(f"  Overall  :  {overall}")
        lines.append('')

    lines.append(THICK)
    lines.append('                    END OF AUDIT REPORT'.center(W))
    lines.append(THICK)
    return '\n'.join(lines)


@app.route('/api/pipeline/start', methods=['POST'])
@login_required
def api_pipeline_start():
    """Probe openFDA then run a multi-step pipeline job in the background.

    pipeline_type values:
      'raw'   - Step 1 only: download raw MAUDE CSV.
      'clean' - Steps 1-2: download + clean, deliver cleaned XLSX.
      'full'  - Steps 1-3: download + clean + IMDRF code counts, deliver XLSX.
    """
    data = request.get_json() or {}
    product_code = (data.get('product_code') or '').strip().upper()
    date_from = (data.get('date_from') or '').strip()
    date_to = (data.get('date_to') or '').strip()

    # Multi-output support: requested_outputs list takes precedence over pipeline_type
    _valid_types = {'raw', 'clean', 'full'}
    _req_raw = data.get('requested_outputs') or []
    if not isinstance(_req_raw, list):
        _req_raw = []
    requested_outputs = [o for o in _req_raw if o in _valid_types]

    if requested_outputs:
        if 'full' in requested_outputs:
            pipeline_type = 'full'
        elif 'clean' in requested_outputs:
            pipeline_type = 'clean'
        else:
            pipeline_type = 'raw'
    else:
        pipeline_type = (data.get('pipeline_type') or 'full').strip().lower()
        if pipeline_type not in _valid_types:
            return jsonify({'error': 'Invalid pipeline_type. Must be raw, clean, or full.'}), 400
        requested_outputs = [pipeline_type]

    if not product_code:
        return jsonify({'error': 'Product code is required.'}), 400

    job_id = str(uuid.uuid4())
    user_id = current_user.id
    safe_code = re.sub(r'[^A-Za-z0-9_-]+', '', product_code.upper()) or 'CODE'

    if pipeline_type == 'raw':
        total_steps = 1
        output_filename = f"maude_{safe_code}_{date_from}_{date_to}.csv"
    elif pipeline_type == 'clean':
        total_steps = 2
        output_filename = f"maude_{safe_code}_{date_from}_{date_to}_cleaned.xlsx"
    else:
        total_steps = 3
        output_filename = f"imdrf_code_counts_{safe_code}_{date_from}_{date_to}.xlsx"

    _set_pipeline_status(
        job_id,
        status='queued', step=0, step_name='Connecting to openFDA…',
        total_steps=total_steps, pipeline_type=pipeline_type,
        processed=0, scanned=0, total=None,
        error=None, output_path=None, output_filename=output_filename,
        requested_outputs=requested_outputs,
        product_code=product_code, date_from=date_from, date_to=date_to,
        user_id=user_id
    )

    def run_pipeline(job_user_id):
        raw_path = os.path.join(app.config['UPLOAD_FOLDER'], f"pipeline_raw_{job_id}.csv")
        cleaned_path = os.path.join(app.config['UPLOAD_FOLDER'], f"pipeline_cleaned_{job_id}.xlsx")
        ext = 'csv' if pipeline_type == 'raw' else 'xlsx'
        final_path = os.path.join(app.config['UPLOAD_FOLDER'], f"pipeline_final_{job_id}.{ext}")

        try:
            # Step 0: Probe openFDA to find correct search field/date combination
            _set_pipeline_status(job_id, status='running', step=0,
                                 step_name='Connecting to openFDA…', user_id=job_user_id)
            err, probe = _maude_probe(product_code, date_from, date_to)
            if err:
                _set_pipeline_status(job_id, status='failed', error=err['error'], user_id=job_user_id)
                return
            _set_pipeline_status(job_id, total=probe.get('total_results'), user_id=job_user_id)

            # Step 1: Download raw MAUDE CSV
            _set_pipeline_status(job_id, status='running', step=1,
                                 step_name='Downloading MAUDE data…', user_id=job_user_id)

            def dl_progress(count, scanned_count):
                _set_pipeline_status(job_id, processed=count, scanned=scanned_count,
                                     user_id=job_user_id)

            step1_dest = final_path if pipeline_type == 'raw' else raw_path

            # For multi-year ranges, fetch each year separately to avoid the
            # openFDA max-skip=25000 cap (which caps total records at ~26,000
            # and, with asc sorting, would return only the earliest year's data).
            try:
                year_from_int = datetime.strptime(date_from, '%Y-%m-%d').year
                year_to_int   = datetime.strptime(date_to,   '%Y-%m-%d').year
            except ValueError:
                year_from_int = year_to_int = None

            with open(step1_dest, 'wb') as f:
                if year_from_int and year_to_int and year_to_int > year_from_int:
                    # Multi-year: fetch each calendar year independently
                    header_written = False
                    for yr in range(year_from_int, year_to_int + 1):
                        y_start = date_from if yr == year_from_int else f"{yr}-01-01"
                        y_end   = date_to   if yr == year_to_int   else f"{yr}-12-31"
                        year_probe = _make_year_probe(probe, y_start, y_end)
                        if year_probe is None:
                            continue  # no data for this year — skip
                        for chunk_idx, chunk in enumerate(_maude_generate_csv(year_probe)):
                            if chunk_idx == 0 and header_written:
                                # Skip the CSV header row for all years after the first
                                continue
                            f.write(chunk)
                            if chunk_idx == 0:
                                header_written = True
                else:
                    # Single year (or unknown range): use original probe directly
                    for chunk in _maude_generate_csv(probe, progress_callback=dl_progress):
                        f.write(chunk)

            # Keep raw file as a secondary output if it was explicitly requested
            if pipeline_type != 'raw' and 'raw' in requested_outputs:
                _set_pipeline_status(job_id, raw_output_path=raw_path, user_id=job_user_id)

            if pipeline_type == 'raw':
                _set_pipeline_status(
                    job_id, status='done', step=1, step_name='Complete',
                    output_path=final_path, output_filename=output_filename,
                    user_id=job_user_id,
                    audit_data={
                        'product_code': product_code,
                        'date_from': date_from,
                        'date_to': date_to,
                        'pipeline_type': pipeline_type,
                        'run_at': datetime.now(timezone.utc).isoformat(),
                        'total_records_found': probe.get('total_results'),
                        'raw_records_downloaded': probe.get('total_results'),
                    },
                )
                return

            # Step 2: Clean the CSV
            _set_pipeline_status(job_id, step=2, step_name='Cleaning data…',
                                 user_id=job_user_id)
            from backend.processor import MAUDEProcessor  # lazy import
            processor = MAUDEProcessor()
            if os.path.exists(DEFAULT_IMDRF_PATH):
                processor.load_imdrf_structure(DEFAULT_IMDRF_PATH)
            stats = processor.process_file(raw_path, cleaned_path)
            if 'raw' not in requested_outputs:
                try:
                    os.remove(raw_path)
                except Exception:
                    pass

            critical_failures = []
            if not stats['validation'].get('column_count_correct', False):
                critical_failures.append('Column count check failed')
            if not stats['validation'].get('file_will_open', False):
                critical_failures.append('File integrity check failed')
            if not stats['validation'].get('date_format_correct', False):
                critical_failures.append('Date format check failed')
            if not stats['validation'].get('imdrf_adjacent', False):
                critical_failures.append('IMDRF Code column position check failed')
            if not stats['validation'].get('imdrf_codes_valid', False):
                critical_failures.append('IMDRF codes validation failed')
            if critical_failures:
                _set_pipeline_status(job_id, status='failed',
                                     error=f"Cleaning validation failed: {', '.join(critical_failures)}",
                                     user_id=job_user_id)
                return

            if pipeline_type == 'clean':
                import shutil
                shutil.move(cleaned_path, final_path)
                _set_pipeline_status(
                    job_id, status='done', step=2, step_name='Complete',
                    output_path=final_path, output_filename=output_filename,
                    user_id=job_user_id,
                    audit_data={
                        'product_code': product_code,
                        'date_from': date_from,
                        'date_to': date_to,
                        'pipeline_type': pipeline_type,
                        'run_at': datetime.now(timezone.utc).isoformat(),
                        'total_records_found': probe.get('total_results'),
                        'raw_records_downloaded': probe.get('total_results'),
                        'stats': {
                            'original_rows': stats.get('original_rows'),
                            'final_rows': stats.get('final_rows'),
                            'rows_removed': stats.get('rows_removed'),
                            'original_cols': stats.get('original_cols'),
                            'final_cols': stats.get('final_cols'),
                            'cols_removed': stats.get('cols_removed', []),
                            'rows_removed_by_reason': stats.get('rows_removed_by_reason', []),
                            'manufacturer_list': stats.get('manufacturer_list', []),
                            'imdrf_stats': stats.get('imdrf_stats', {}),
                            'validation': stats.get('validation', {}),
                        },
                    },
                )
                return

            # Step 3: Generate IMDRF code counts XLSX
            _set_pipeline_status(job_id, step=3, step_name='Generating code counts…',
                                 user_id=job_user_id)
            from backend.imdrf_insights import (  # lazy import
                _load_cleaned_dataframe,
                get_imdrf_code_counts_all_levels_with_descriptions,
                get_imdrf_code_monthly_counts,
                get_patient_problem_e_code_monthly_counts,
                load_imdrf_code_descriptions,
                EXCLUDED_IMDRF_L1_PREFIXES,
            )
            from openpyxl import Workbook
            from openpyxl.styles import Font

            cleaned_df = _load_cleaned_dataframe(cleaned_path)
            counts_by_level = get_imdrf_code_counts_all_levels_with_descriptions(
                cleaned_path, DEFAULT_IMDRF_PATH, df=cleaned_df)
            monthly_data = get_imdrf_code_monthly_counts(cleaned_path, df=cleaned_df)
            patient_e_monthly = get_patient_problem_e_code_monthly_counts(
                cleaned_path, DEFAULT_IMDRF_PATH, df=cleaned_df)

            # ── Audit: per-level code count summary ──────────────────────────
            level_summary = {}
            for _lvl in [1, 2, 3]:
                _lvl_data = counts_by_level.get(_lvl, {})
                level_summary[str(_lvl)] = {
                    'distinct_codes': len(_lvl_data),
                    'total_instances': sum(d.get('count', 0) for d in _lvl_data.values()),
                }

            # ── Audit: A24/A25 excluded rows (from cleaned MAUDE data) ───────
            _a24_a25_row_count = 0
            _imdrf_code_col = next(
                (c for c in cleaned_df.columns if c == 'IMDRF Code'), None
            )
            if _imdrf_code_col:
                _a24_a25_row_count = int(
                    cleaned_df[_imdrf_code_col]
                    .astype(str).str[:3].str.upper()
                    .isin(EXCLUDED_IMDRF_L1_PREFIXES)
                    .sum()
                )

            # ── Audit: A24/A25 descriptions from Annex ───────────────────────
            _a24_a25_descriptions = {}
            try:
                if os.path.exists(DEFAULT_IMDRF_PATH):
                    _desc_map = load_imdrf_code_descriptions(DEFAULT_IMDRF_PATH)
                    for _code in sorted(EXCLUDED_IMDRF_L1_PREFIXES):
                        _a24_a25_descriptions[_code] = (
                            _desc_map.get(1, {}).get(_code, 'Description not available')
                        )
            except Exception:
                pass

            if 'clean' in requested_outputs:
                _set_pipeline_status(job_id, clean_output_path=cleaned_path, user_id=job_user_id)
            else:
                try:
                    os.remove(cleaned_path)
                except Exception:
                    pass

            wb = Workbook()
            ws = wb.active
            ws.title = "IMDRF Code Counts"
            bold_font = Font(bold=True)

            all_months = monthly_data.get('months', [])
            monthly_counts = monthly_data.get('counts', {})
            num_months = len(all_months)

            # Format 'YYYY-MM' → 'MMM-YYYY' for column headers
            def _fmt_month(ym):
                try:
                    import calendar
                    y, m = ym.split('-')
                    return f"{calendar.month_abbr[int(m)]}-{y}"
                except Exception:
                    return ym

            month_headers = [_fmt_month(m) for m in all_months]

            for level in [1, 2, 3]:
                level_label = f"LEVEL-{level} Code"
                ws.append([level_label] + [""] * (3 + num_months))
                ws.cell(row=ws.max_row, column=1).font = bold_font

                header_row = ["IMDRF Code", "Description", "Total", "Avg/Month"] + month_headers
                ws.append(header_row)
                for col_idx in range(1, len(header_row) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).font = bold_font

                level_counts = counts_by_level.get(level, {})
                level_monthly = monthly_counts.get(level, {})
                for code in sorted(level_counts.keys(), key=str):
                    row_data = level_counts.get(code, {})
                    total = row_data.get('count', 0)
                    avg = round(total / num_months, 2) if num_months > 0 else 0
                    code_monthly = level_monthly.get(str(code), {})
                    month_vals = [code_monthly.get(m, 0) for m in all_months]
                    ws.append([str(code), row_data.get('description', ''), total, avg] + month_vals)
                ws.append([""] * (4 + num_months))

            # ── Patient Problem E-code sections (L1 / L2 / L3) ──────────────
            pat_months = patient_e_monthly.get('months', [])
            pat_monthly_counts = patient_e_monthly.get('counts', {})
            pat_totals = patient_e_monthly.get('totals', {})
            num_pat_months = len(pat_months)
            pat_month_headers = [_fmt_month(m) for m in pat_months]

            for level in [1, 2, 3]:
                level_label = f"PATIENT PROBLEM E-CODES LEVEL-{level}"
                ws.append([level_label] + [""] * (3 + num_pat_months))
                ws.cell(row=ws.max_row, column=1).font = bold_font

                header_row = ["IMDRF E-Code", "Description", "Total", "Avg/Month"] + pat_month_headers
                ws.append(header_row)
                for col_idx in range(1, len(header_row) + 1):
                    ws.cell(row=ws.max_row, column=col_idx).font = bold_font

                level_totals = pat_totals.get(level, {})
                level_monthly = pat_monthly_counts.get(level, {})
                for code in sorted(level_totals.keys(), key=str):
                    row_data = level_totals[code]
                    total = row_data.get('count', 0)
                    avg = round(total / num_pat_months, 2) if num_pat_months > 0 else 0
                    code_monthly = level_monthly.get(str(code), {})
                    month_vals = [code_monthly.get(m, 0) for m in pat_months]
                    ws.append([str(code), row_data.get('description', ''), total, avg] + month_vals)
                ws.append([""] * (4 + num_pat_months))

            wb.save(final_path)
            _set_pipeline_status(
                job_id, status='done', step=3, step_name='Complete',
                output_path=final_path, output_filename=output_filename,
                user_id=job_user_id,
                audit_data={
                    'product_code': product_code,
                    'date_from': date_from,
                    'date_to': date_to,
                    'pipeline_type': pipeline_type,
                    'run_at': datetime.now(timezone.utc).isoformat(),
                    'total_records_found': probe.get('total_results'),
                    'raw_records_downloaded': probe.get('total_results'),
                    'stats': {
                        'original_rows': stats.get('original_rows'),
                        'final_rows': stats.get('final_rows'),
                        'rows_removed': stats.get('rows_removed'),
                        'original_cols': stats.get('original_cols'),
                        'final_cols': stats.get('final_cols'),
                        'cols_removed': stats.get('cols_removed', []),
                        'rows_removed_by_reason': stats.get('rows_removed_by_reason', []),
                        'manufacturer_list': stats.get('manufacturer_list', []),
                        'imdrf_stats': stats.get('imdrf_stats', {}),
                        'validation': stats.get('validation', {}),
                    },
                    'level_summary': level_summary,
                    'a24_a25_excluded': _a24_a25_descriptions,
                    'a24_a25_row_count': _a24_a25_row_count,
                },
            )

        except Exception as e:
            _set_pipeline_status(job_id, status='failed', error=str(e), user_id=job_user_id)
        finally:
            # Only clean up intermediate files that were NOT explicitly requested as outputs
            for p, ftype in [(raw_path, 'raw'), (cleaned_path, 'clean')]:
                if ftype not in requested_outputs and os.path.exists(p):
                    try:
                        os.remove(p)
                    except Exception:
                        pass

    thread = threading.Thread(target=run_pipeline, args=(user_id,), daemon=True)
    thread.start()
    return jsonify({
        'job_id': job_id,
        'status_url': url_for('api_pipeline_status', job_id=job_id),
        'download_url': url_for('api_pipeline_download', job_id=job_id),
        'audit_url': url_for('api_pipeline_audit', job_id=job_id),
    }), 202


@app.route('/api/pipeline/status/<job_id>', methods=['GET'])
@login_required
def api_pipeline_status(job_id):
    job = _get_pipeline_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    return jsonify({
        'status': job.get('status'),
        'step': job.get('step'),
        'step_name': job.get('step_name'),
        'total_steps': job.get('total_steps'),
        'pipeline_type': job.get('pipeline_type'),
        'processed': job.get('processed'),
        'total': job.get('total'),
        'scanned': job.get('scanned'),
        'error': job.get('error'),
    }), 200


@app.route('/api/pipeline/download/<job_id>', methods=['GET'])
@login_required
def api_pipeline_download(job_id):
    job = _get_pipeline_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'File not ready yet'}), 409
    output_path = job.get('output_path')
    output_filename = job.get('output_filename', 'download.csv')
    if not output_path or not os.path.exists(output_path):
        return jsonify({'error': 'Output file not found'}), 404
    pipeline_type = job.get('pipeline_type', 'raw')
    if pipeline_type == 'raw':
        mimetype = 'text/csv'
    else:
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return send_file(output_path, as_attachment=True, download_name=output_filename, mimetype=mimetype)


@app.route('/api/pipeline/audit/<job_id>', methods=['GET'])
@login_required
def api_pipeline_audit(job_id):
    """Return a plain-text audit report for a completed pipeline job."""
    job = _get_pipeline_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'Pipeline must be complete before downloading the audit report'}), 409

    audit = job.get('audit_data') or {}
    report_text = _build_audit_report(audit)

    safe_code = re.sub(r'[^A-Za-z0-9_-]+', '', (audit.get('product_code') or 'CODE').upper())
    filename = f"audit_report_{safe_code}_{audit.get('date_from', '')}_{audit.get('date_to', '')}.txt"

    output = io.BytesIO(report_text.encode('utf-8'))
    output.seek(0)
    return send_file(output, mimetype='text/plain', as_attachment=True, download_name=filename)


@app.route('/api/pipeline/download-file/<job_id>/<file_type>', methods=['GET'])
@login_required
def api_pipeline_download_file(job_id, file_type):
    """Download a specific output file (raw/clean/full) from a completed pipeline job."""
    if file_type not in ('raw', 'clean', 'full'):
        return jsonify({'error': 'Invalid file type. Must be raw, clean, or full.'}), 400
    job = _get_pipeline_job(job_id)
    if not job or job.get('user_id') != current_user.id:
        return jsonify({'error': 'Job not found'}), 404
    if job.get('status') != 'done':
        return jsonify({'error': 'Job not complete yet'}), 409

    safe_code = re.sub(r'[^A-Za-z0-9_-]+', '', (job.get('product_code') or 'CODE').upper()) or 'CODE'
    date_from = job.get('date_from', '')
    date_to = job.get('date_to', '')

    upload_folder = app.config['UPLOAD_FOLDER']

    if file_type == 'raw':
        # Stored path → fallback to expected temp filename for this job
        path = job.get('raw_output_path')
        if not path or not os.path.exists(path):
            path = job.get('output_path') if job.get('pipeline_type') == 'raw' else None
        if not path or not os.path.exists(path):
            path = os.path.join(upload_folder, f"pipeline_raw_{job_id}.csv")
        filename = f"maude_{safe_code}_{date_from}_{date_to}.csv"
        mimetype = 'text/csv'
    elif file_type == 'clean':
        path = job.get('clean_output_path')
        if not path or not os.path.exists(path):
            path = job.get('output_path') if job.get('pipeline_type') == 'clean' else None
        if not path or not os.path.exists(path):
            path = os.path.join(upload_folder, f"pipeline_cleaned_{job_id}.xlsx")
        filename = f"maude_{safe_code}_{date_from}_{date_to}_cleaned.xlsx"
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    else:  # full
        path = job.get('output_path') if job.get('pipeline_type') == 'full' else None
        if not path or not os.path.exists(path):
            path = os.path.join(upload_folder, f"pipeline_final_{job_id}.xlsx")
        filename = job.get('output_filename', f"imdrf_code_counts_{safe_code}_{date_from}_{date_to}.xlsx")
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    if not path or not os.path.exists(path):
        requested = job.get('requested_outputs', [])
        return jsonify({
            'error': f'"{file_type}" file was not produced by this job. '
                     f'Pipeline type was "{job.get("pipeline_type")}", '
                     f'requested outputs were {requested}.'
        }), 404
    return send_file(path, as_attachment=True, download_name=filename, mimetype=mimetype)


# Device Recall Routes
@app.route('/device-recall')
@login_required
def device_recall_page():
    """Render Device Recall Search page."""
    return render_template('device_recall.html', user=current_user)


@app.route('/api/device-recall/search', methods=['POST'])
@login_required
def api_device_recall_search():
    """Search FDA device recalls via openFDA /device/recall.json."""
    data = request.get_json() or {}
    product_code = (data.get('product_code') or '').strip().upper()
    date_from = (data.get('date_from') or '').strip()
    date_to = (data.get('date_to') or '').strip()

    if not product_code:
        return jsonify({'error': 'Product code is required.'}), 400
    if not date_from or not date_to:
        return jsonify({'error': 'Date range is required.'}), 400

    try:
        start_date = datetime.strptime(date_from, '%Y-%m-%d')
        end_date   = datetime.strptime(date_to,   '%Y-%m-%d')
    except ValueError:
        return jsonify({'error': 'Invalid date format. Use YYYY-MM-DD.'}), 400

    if start_date > end_date:
        return jsonify({'error': 'Date From must be on or before Date To.'}), 400

    start_openfda = start_date.strftime('%Y%m%d')
    end_openfda   = end_date.strftime('%Y%m%d')
    base_url = 'https://api.fda.gov/device/recall.json'

    # Try product_code field first, then openfda.product_code
    search_candidates = [
        f'product_code:{_format_openfda_search_value(product_code)} AND recall_initiation_date:[{start_openfda} TO {end_openfda}]',
        f'openfda.product_code:{_format_openfda_search_value(product_code)} AND recall_initiation_date:[{start_openfda} TO {end_openfda}]',
    ]

    records = []
    total = 0
    found = False
    for search_str in search_candidates:
        try:
            resp = _openfda_get(base_url, params={'search': search_str, 'limit': 100},
                                allow_not_found=True, allow_bad_request=True)
            if resp.status_code == 200:
                payload = resp.json()
                results = payload.get('results') or []
                if results:
                    records = results
                    total = payload.get('meta', {}).get('results', {}).get('total', len(results))
                    found = True
                    break
        except Exception:
            continue

    if not found:
        return jsonify({'error': f'No recall records found for product code {product_code} in the specified date range.'}), 404

    def _safe(val):
        if val is None:
            return ''
        if isinstance(val, list):
            return '; '.join(str(v) for v in val if v)
        return str(val)

    rows = []
    for r in records:
        openfda = r.get('openfda') or {}
        rows.append({
            'recall_number':         _safe(r.get('recall_number')),
            'recalling_firm':        _safe(r.get('recalling_firm')),
            'product_description':   _safe(r.get('product_description')),
            'reason_for_recall':     _safe(r.get('reason_for_recall')),
            'status':                _safe(r.get('status')),
            'classification':        _safe(r.get('classification')),
            'voluntary_mandated':    _safe(r.get('voluntary_mandated')),
            'recall_initiation_date':_safe(r.get('recall_initiation_date')),
            'center_classification_date': _safe(r.get('center_classification_date')),
            'distribution_pattern':  _safe(r.get('distribution_pattern')),
            'product_quantity':      _safe(r.get('product_quantity')),
            'product_code':          _safe(openfda.get('product_code') or r.get('product_code')),
            'device_name':           _safe(openfda.get('device_name')),
        })

    return jsonify({'success': True, 'total': total, 'returned': len(rows), 'records': rows}), 200


# TXT to CSV Converter Routes
@app.route('/txt-to-csv')
@login_required
def txt_to_csv_page():
    """Render TXT to CSV converter page."""
    return render_template('txt_to_csv.html', user=current_user)


@app.route('/api/txt-to-csv/upload-chunk', methods=['POST'])
@login_required
def api_txt_upload_chunk():
    """Handle chunked file upload for large TXT files (10GB+ support)."""
    try:
        chunk = request.files.get('chunk')
        chunk_index = int(request.form.get('chunkIndex', 0))
        total_chunks = int(request.form.get('totalChunks', 1))
        file_id = request.form.get('fileId')
        original_filename = request.form.get('filename', 'upload.txt')

        if not chunk:
            return jsonify({'error': 'No chunk data received'}), 400

        # Create file_id on first chunk
        if chunk_index == 0:
            file_id = f"{current_user.id}_{os.urandom(8).hex()}"

        if not file_id:
            return jsonify({'error': 'Missing file ID'}), 400

        # Secure the filename
        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

        # Append chunk to file (create on first chunk, append on subsequent)
        mode = 'wb' if chunk_index == 0 else 'ab'
        with open(file_path, mode) as f:
            # Stream the chunk in smaller pieces to avoid memory issues
            while True:
                data = chunk.read(8192)  # 8KB at a time
                if not data:
                    break
                f.write(data)

        # Return file_id and status
        is_complete = (chunk_index + 1) >= total_chunks

        return jsonify({
            'success': True,
            'fileId': file_id,
            'chunkIndex': chunk_index,
            'isComplete': is_complete
        }), 200

    except Exception as e:
        return jsonify({'error': f'Chunk upload failed: {str(e)}'}), 400


@app.route('/api/txt-to-csv/preview', methods=['POST'])
@login_required
def api_txt_preview():
    """Get preview for already uploaded TXT file (by file_id) or upload small file."""
    try:
        # Check if this is a file_id based preview request (for chunked uploads)
        if request.is_json:
            data = request.get_json()
            file_id = data.get('file_id')
            original_filename = data.get('filename', 'upload.txt')

            if file_id:
                filename = secure_filename(original_filename)
                file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

                if not os.path.exists(file_path):
                    return jsonify({'error': 'File not found. Please upload again.'}), 404

                # Get preview
                preview = get_txt_preview(file_path, num_rows=10)

                # Store file path in session
                session[f'txt_file_{file_id}'] = file_path

                return jsonify({
                    'success': True,
                    'file_id': file_id,
                    'preview': preview
                }), 200

        # Fallback: Handle traditional file upload for smaller files
        if 'file' not in request.files:
            return jsonify({'error': 'No file uploaded'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400

        if not file.filename.lower().endswith('.txt'):
            return jsonify({'error': 'Invalid file type. Please upload a .txt file.'}), 400

        # Save uploaded file with unique name
        filename = secure_filename(file.filename)
        file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        input_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{file_id}_{filename}")

        # Stream file to disk in chunks
        with open(input_path, 'wb') as f:
            while True:
                chunk = file.read(8192)  # 8KB chunks
                if not chunk:
                    break
                f.write(chunk)

        # Get preview
        preview = get_txt_preview(input_path, num_rows=10)

        # Store file path in session
        session[f'txt_file_{file_id}'] = input_path

        return jsonify({
            'success': True,
            'file_id': file_id,
            'preview': preview
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/txt-to-csv/convert', methods=['POST'])
@login_required
def api_txt_convert():
    """Convert TXT file to CSV."""
    data = request.get_json()
    file_id = data.get('file_id')

    if not file_id:
        return jsonify({'error': 'Missing file_id parameter'}), 400

    try:
        # Retrieve file path from session
        file_path = session.get(f'txt_file_{file_id}')
        if not file_path or not os.path.exists(file_path):
            return jsonify({'error': 'File not found. Please upload again.'}), 404

        # Generate output path
        output_file_id = f"{current_user.id}_{os.urandom(8).hex()}"
        output_filename = os.path.splitext(os.path.basename(file_path))[0] + '.csv'
        output_path = os.path.join(app.config['UPLOAD_FOLDER'], f"{output_file_id}_{output_filename}")

        # Convert file
        converter = TxtToCsvConverter()
        stats = converter.process_file_chunked(file_path, output_path)

        # Store output path in session
        session[f'csv_file_{output_file_id}'] = {
            'path': output_path,
            'filename': output_filename
        }

        return jsonify({
            'success': True,
            'output_file_id': output_file_id,
            'stats': stats
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/txt-to-csv/download/<file_id>')
@login_required
def api_txt_download(file_id):
    """Download converted CSV file."""
    try:
        # Retrieve file info from session
        file_info = session.get(f'csv_file_{file_id}')
        if not file_info or not os.path.exists(file_info['path']):
            return jsonify({'error': 'File not found. Please convert again.'}), 404

        return send_file(
            file_info['path'],
            as_attachment=True,
            download_name=file_info['filename'],
            mimetype='text/csv'
        )

    except Exception as e:
        return jsonify({'error': str(e)}), 400


# CSV Viewer Routes
@app.route('/csv-viewer')
@login_required
def csv_viewer_page():
    """Render CSV viewer page."""
    return render_template('csv_viewer.html', user=current_user)


@app.route('/api/csv-viewer/upload-chunk', methods=['POST'])
@login_required
def api_csv_upload_chunk():
    """Handle chunked file upload for CSV viewer."""
    try:
        chunk = request.files.get('chunk')
        chunk_index = int(request.form.get('chunkIndex', 0))
        total_chunks = int(request.form.get('totalChunks', 1))
        file_id = request.form.get('fileId')
        original_filename = request.form.get('filename', 'upload.csv')

        if not chunk:
            return jsonify({'error': 'No chunk data received'}), 400

        # Create file_id on first chunk
        if chunk_index == 0:
            file_id = f"{current_user.id}_{os.urandom(8).hex()}"

        if not file_id:
            return jsonify({'error': 'Missing file ID'}), 400

        # Secure the filename
        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        # Append chunk to file
        mode = 'wb' if chunk_index == 0 else 'ab'
        with open(file_path, mode) as f:
            while True:
                data = chunk.read(8192)
                if not data:
                    break
                f.write(data)

        return jsonify({
            'success': True,
            'fileId': file_id,
            'chunkIndex': chunk_index,
            'isComplete': (chunk_index + 1) >= total_chunks
        }), 200

    except Exception as e:
        return jsonify({'error': f'Chunk upload failed: {str(e)}'}), 400


@app.route('/api/csv-viewer/info', methods=['POST'])
@login_required
def api_csv_info():
    """Get CSV file information."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Store in session
        session[f'csv_viewer_{file_id}'] = file_path

        # Get file info
        info = get_csv_info(file_path)

        return jsonify({
            'success': True,
            'info': info
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/page', methods=['POST'])
@login_required
def api_csv_page():
    """Get a page of CSV data."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        page = data.get('page', 1)
        page_size = data.get('page_size', 100)

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Get page data
        page_data = get_csv_page(file_path, page, page_size)

        return jsonify(page_data), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/search', methods=['POST'])
@login_required
def api_csv_search():
    """Search in CSV file."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        search_term = data.get('search_term', '')

        if not file_id or not search_term:
            return jsonify({'error': 'Missing file_id or search_term'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Search in file
        viewer = LargeCSVViewer()
        results = viewer.search_in_file(file_path, search_term)

        return jsonify(results), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


@app.route('/api/csv-viewer/column-stats', methods=['POST'])
@login_required
def api_csv_column_stats():
    """Get column statistics."""
    try:
        data = request.get_json()
        file_id = data.get('file_id')
        original_filename = data.get('filename', 'upload.csv')
        column_index = data.get('column_index', 0)

        if not file_id:
            return jsonify({'error': 'Missing file_id'}), 400

        filename = secure_filename(original_filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], f"csv_{file_id}_{filename}")

        if not os.path.exists(file_path):
            return jsonify({'error': 'File not found'}), 404

        # Get column stats
        viewer = LargeCSVViewer()
        stats = viewer.get_column_stats(file_path, column_index)

        return jsonify({
            'success': True,
            'stats': stats
        }), 200

    except Exception as e:
        return jsonify({'error': str(e)}), 400


if __name__ == '__main__':
    # Check for Groq API key
    if not GROQ_API_KEY:
        print("WARNING: GROQ_API_KEY not set. Set it as an environment variable.")

    app.run(debug=True, host='0.0.0.0', port=5000)
